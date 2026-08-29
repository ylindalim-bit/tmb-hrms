# -*- coding: utf-8 -*-
import calendar
import datetime
import functools
import io
import json
import os
import secrets
import sqlite3
import uuid
import zipfile

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins


def _set_a4_one_page(ws):
    """Configures a worksheet's print setup so it fits on one A4 page:
    landscape, shrunk to 1 page wide x 1 page tall, narrow margins."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True


def _add_headcount_block(ws, start_row, totals, employee_count):
    """Head count + statutory contribution (Employee/Employer/Total) summary
    box, written starting at start_row in columns A-E. Shared by both
    payroll export routes so the block stays identical in each.
    SOCSO/SKBBK/EIS are grouped with a brace and combined total (all three
    go to PERKESO in one combined submission - see the SOCSO+EIS text
    file); EPF/PCB/HRDCORP each go to a different body, so they're listed
    separately with no combined figure."""
    hc_row = start_row
    ws.cell(row=hc_row, column=1, value="Head count").font = Font(bold=True)
    ws.cell(row=hc_row, column=2, value=f"{employee_count} employees")

    header_row = hc_row + 2
    for col_idx, label in [(3, "E'yee"), (4, "E'yer"), (5, "Total")]:
        ws.cell(row=header_row, column=col_idx, value=label).font = Font(bold=True)

    perkeso_rows = [
        ("SOCSO", totals["socso_employee"], totals["socso_employer"]),
        ("SKBBK", totals["skbbk_employee"], None),
        ("EIS", totals["eis_employee"], totals["eis_employer"]),
    ]
    other_rows = [
        ("EPF", totals["epf_employee"], totals["epf_employer"]),
        ("PCB", totals["pcb"], None),
        ("HRDCORP", None, totals["hrd_levy_employer"]),
    ]

    def _write_row(r_idx, label, eyee, eyer):
        ws.cell(row=r_idx, column=1, value=label).font = Font(color="1D4ED8")
        eyee_val = eyee if eyee is not None else 0
        eyer_val = eyer if eyer is not None else 0
        for col_idx, val in [(3, eyee_val), (4, eyer_val), (5, round(eyee_val + eyer_val, 2))]:
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.number_format = "#,##0.00"
        return round(eyee_val + eyer_val, 2)

    r_idx = header_row + 1
    perkeso_first_row = r_idx
    perkeso_total = 0
    for label, eyee, eyer in perkeso_rows:
        perkeso_total += _write_row(r_idx, label, eyee, eyer)
        r_idx += 1
    perkeso_last_row = r_idx - 1

    # Brace (column G) + combined PERKESO total (column H), vertically
    # centered across the SOCSO/SKBBK/EIS rows.
    brace_cell = ws.cell(row=perkeso_first_row, column=7, value="}")
    brace_cell.font = Font(size=28)
    brace_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=perkeso_first_row, start_column=7, end_row=perkeso_last_row, end_column=7)
    total_cell = ws.cell(row=perkeso_first_row, column=8, value=round(perkeso_total, 2))
    total_cell.font = Font(bold=True)
    total_cell.number_format = "#,##0.00"
    total_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=perkeso_first_row, start_column=8, end_row=perkeso_last_row, end_column=8)

    r_idx += 1  # blank row between groups
    for label, eyee, eyer in other_rows:
        _write_row(r_idx, label, eyee, eyer)
        r_idx += 1

    box_last_row = r_idx - 1
    box_last_col = 5
    thin = Side(style="thin", color="1F7A3D")
    for row in ws.iter_rows(min_row=hc_row, max_row=box_last_row, min_col=1, max_col=box_last_col):
        for cell in row:
            top = thin if cell.row == hc_row else None
            bottom = thin if cell.row == box_last_row else None
            left = thin if cell.column == 1 else None
            right = thin if cell.column == box_last_col else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    return box_last_row


def _add_zero_pay_notes(ws, start_row, notes):
    """Writes each zero-pay explanation on its own row starting at
    start_row, in a highlighted amber font. Returns the last row used (or
    start_row - 1 if there were no notes)."""
    if not notes:
        return start_row - 1
    ws.cell(row=start_row, column=1, value="Note:").font = Font(bold=True, color="92400E")
    row = start_row
    for note in notes:
        ws.cell(row=row, column=1, value=f"- {note}").font = Font(color="92400E")
        row += 1
    return row - 1


from flask import Flask, Response, abort, g, redirect, render_template, request, send_from_directory, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

import payroll_calc

# DATA_DIR points at wherever the real data should live. Locally this is
# just the app folder (unchanged behavior); in production (Render) it's set
# via the DATA_DIR env var to the mounted Persistent Disk, so the database,
# uploads, and secret key all survive restarts/redeploys instead of living
# on the container's throwaway filesystem.
DATA_DIR = os.environ.get("DATA_DIR", os.path.dirname(__file__))
DB_PATH = os.path.join(DATA_DIR, "payroll.db")
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")

if not os.path.exists(DB_PATH):
    # First run on a fresh environment (e.g. a brand new Render deploy) -
    # build an empty database from schema.sql + seed_reference_data.sql
    # (statutory rate tables only, no employee/payroll data) so the app has
    # something to connect to instead of crashing.
    os.makedirs(DATA_DIR, exist_ok=True)
    _bootstrap_conn = sqlite3.connect(DB_PATH)
    _here = os.path.dirname(__file__)
    with open(os.path.join(_here, "schema.sql"), "r", encoding="utf-8") as _f:
        _bootstrap_conn.executescript(_f.read())
    _seed_path = os.path.join(_here, "seed_reference_data.sql")
    if os.path.exists(_seed_path):
        with open(_seed_path, "r", encoding="utf-8") as _f:
            _bootstrap_conn.executescript(_f.read())
    _bootstrap_conn.commit()
    _bootstrap_conn.close()
DOCUMENT_TYPES = ["Job Application Form", "IC / Passport Copy", "Letter of Employment", "Confirmation Letter",
                   "Resignation Letter", "CP22A", "e-Stamping Certificate", "TP3 (Prior Employer Income)", "Other"]
BUSINESS_TRIP_TYPES = ["Business Trip", "Out-Duty", "Training", "Unrecorded Leave"]
ALLOWED_DOC_EXTENSIONS = {"pdf", "doc", "docx", "jpg", "jpeg", "png"}
ALLOWED_PHOTO_EXTENSIONS = {"jpg", "jpeg", "png"}
RACE_OPTIONS = ["Malay", "Chinese", "Iban", "Kadazan", "Other"]
RELIGION_OPTIONS = ["Islam", "Buddha", "Christian", "Other"]
HOLIDAY_STATE_OPTIONS = ["Johor", "China"]
MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]

PAYMENT_DAY = 7  # payroll for a month is paid on this day of the following month


def payment_date_for(year, month):
    pay_year, pay_month = (year + 1, 1) if month == 12 else (year, month + 1)
    return datetime.date(pay_year, pay_month, PAYMENT_DAY)


def payslip_release_date(db, year, month):
    """The date this month's payslip becomes visible to the employee on the
    Staff Portal - a day-of-month setting HR controls (Settings page),
    independent of Payment Day so a payslip can be shown a few days before
    or after actual payment if desired. Defaults to the 7th of the
    following month if never configured."""
    row = db.execute("SELECT value FROM payroll_settings WHERE key='payslip_release_day'").fetchone()
    release_day = int(row["value"]) if row else PAYMENT_DAY
    pay_year, pay_month = (year + 1, 1) if month == 12 else (year, month + 1)
    days_in_release_month = calendar.monthrange(pay_year, pay_month)[1]
    return datetime.date(pay_year, pay_month, min(release_day, days_in_release_month))

app = Flask(__name__)
# Persisted to a file (under DATA_DIR - the Persistent Disk in production)
# rather than regenerated per process start: the server restarts often
# (crashes, Windows Startup, active development, a Render redeploy), and a
# new random key every time would silently log every logged-in employee out
# of the Staff Portal on every restart, not just after a real security event.
SECRET_KEY_PATH = os.path.join(DATA_DIR, ".secret_key")
if os.path.exists(SECRET_KEY_PATH):
    with open(SECRET_KEY_PATH, "r") as f:
        app.secret_key = f.read().strip()
else:
    app.secret_key = secrets.token_hex(32)
    with open(SECRET_KEY_PATH, "w") as f:
        f.write(app.secret_key)


@app.context_processor
def inject_pending_counts():
    # Available on every HR-side page via base.html's nav, so a pending Leave
    # Request or Business Trip is visible at a glance instead of only showing
    # up after clicking into that page - the closest thing to a notification
    # this local-only app (no email/SMS setup) can offer.
    try:
        db = get_db()
        if session.get("hr_role") == "approver":
            # Scoped to just the employees assigned to this approver (e.g. Mr
            # Kee), matching what leave_requests_admin() itself will show them.
            pending_leave_count = db.execute(
                """SELECT COUNT(*) AS c FROM leave_requests lr
                   JOIN employees e ON e.emp_id = lr.emp_id
                   WHERE lr.status='Pending' AND e.leave_approver_username=?""",
                (session["hr_username"],),
            ).fetchone()["c"]
            pending_trip_count = db.execute(
                """SELECT COUNT(*) AS c FROM business_trips bt
                   JOIN employees e ON e.emp_id = bt.emp_id
                   WHERE bt.status='Pending' AND e.leave_approver_username=?""",
                (session["hr_username"],),
            ).fetchone()["c"]
            pending_medical_claim_count = 0  # Medical Claims is outside an approver's restricted access
            # For an approver (e.g. Mr Kee), the reminder is "how many of my
            # team have never had an appraisal at all" - there's no due-date
            # field to compare against, so "never appraised" is the signal.
            pending_appraisal_count = db.execute(
                """SELECT COUNT(*) AS c FROM employees e
                   WHERE e.appraisal_supervisor_username=? AND e.status != 'Inactive'
                     AND NOT EXISTS (SELECT 1 FROM appraisals a WHERE a.emp_id = e.emp_id)""",
                (session["hr_username"],),
            ).fetchone()["c"]
        else:
            pending_leave_count = db.execute(
                "SELECT COUNT(*) AS c FROM leave_requests WHERE status='Pending'"
            ).fetchone()["c"]
            pending_trip_count = db.execute(
                "SELECT COUNT(*) AS c FROM business_trips WHERE status='Pending'"
            ).fetchone()["c"]
            pending_medical_claim_count = db.execute(
                "SELECT COUNT(*) AS c FROM medical_claims WHERE status='Pending'"
            ).fetchone()["c"]
            pending_appraisal_count = db.execute(
                "SELECT COUNT(*) AS c FROM appraisals WHERE status='Submitted' AND hr_viewed_at IS NULL"
            ).fetchone()["c"]
        pending_profile_update_count = db.execute(
            "SELECT COUNT(*) AS c FROM profile_update_log WHERE hr_viewed_at IS NULL"
        ).fetchone()["c"] if session.get("hr_role") != "approver" else 0
        # can_approve_ot is a single company-wide approver (not a
        # per-employee assignment), so this is the same total for both an
        # approver and full HR admin - no scoping needed.
        pending_ot_claim_count = db.execute(
            "SELECT COUNT(*) AS c FROM ot_claims WHERE status='Pending'"
        ).fetchone()["c"]
    except sqlite3.OperationalError:
        return {}
    return {"pending_leave_count": pending_leave_count, "pending_trip_count": pending_trip_count,
            "pending_medical_claim_count": pending_medical_claim_count,
            "pending_appraisal_count": pending_appraisal_count,
            "pending_profile_update_count": pending_profile_update_count,
            "pending_ot_claim_count": pending_ot_claim_count}


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


# Paths reachable without an HR login. Deny-by-default (checked in
# require_hr_login below) rather than a decorator on every individual HR
# route, since a dozens-of-routes app makes it too easy to forget one.
HR_LOGIN_EXEMPT_PREFIXES = (
    "/static/",
    "/portal",       # Staff Portal has its own separate @portal_login_required gate
    "/api/",         # read-only payroll-export feed consumed by another local process
    "/hr/login",
    "/hr/logout",
    "/hr/setup",     # first-run only - locks itself once any hr_users row exists
    "/hr/restore-database",  # gated by RESTORE_TOKEN env var, not session - see route
    "/hr/restore-uploads",   # gated by RESTORE_TOKEN env var, not session - see route
    "/hr/migrate-schema",    # gated by RESTORE_TOKEN env var, not session - see route
    "/hr/import-historical-payroll",  # gated by RESTORE_TOKEN env var, not session - see route
    "/hr/bulk-set-medical-claim-limit",  # gated by RESTORE_TOKEN env var, not session - see route
    "/hr/backup-database",   # gated by RESTORE_TOKEN env var, not session - see route
    "/hr/backup-uploads",    # gated by RESTORE_TOKEN env var, not session - see route
    "/hr/seed-attendance-daily",  # gated by RESTORE_TOKEN env var, not session - see route
    "/hr/adjust-attendance",      # gated by RESTORE_TOKEN env var, not session - see route
    "/hr/seed-ot-claims",         # gated by RESTORE_TOKEN env var, not session - see route
    "/hr/ot-claims-cleanup",      # gated by RESTORE_TOKEN env var, not session - see route
)

# role='approver' users (e.g. Mr Kee) get a restricted account: leave
# requests only, plus the housekeeping routes every logged-in user needs.
# Which specific requests they can act on is enforced separately in
# leave_requests_admin/review_leave_request via employees.leave_approver_username.
# Always allowed for a logged-in approver, regardless of which capability
# flags they have.
APPROVER_ALWAYS_ALLOWED_PREFIXES = ("/hr/logout", "/hr/change-password")
# Capability flag (session key) -> path prefixes it unlocks for a
# role='approver' account. Mr Kee has both leave and appraisal; Mr Yang has
# appraisal and (now) OT claims.
APPROVER_CAPABILITY_PREFIXES = {
    # Business Trips reuses the same Leave Approver assignment - it's the
    # same "who's this person's supervisor" relationship, not a separate one.
    "can_approve_leave": ("/leave-requests", "/business-trips"),
    "can_approve_appraisal": ("/appraisals",),
    "can_approve_ot": ("/ot-claims",),
}


def _approver_home():
    """Where to send a role='approver' account after login or when they hit
    a page outside their capabilities - whichever section their flags grant,
    preferring Leave Requests if they have several."""
    if session.get("can_approve_leave") == "Y":
        return url_for("leave_requests_admin")
    if session.get("can_approve_appraisal") == "Y":
        return url_for("appraisal_team")
    if session.get("can_approve_ot") == "Y":
        return url_for("ot_claims_admin")
    return url_for("hr_change_password")


@app.before_request
def require_hr_login():
    if request.path.startswith(HR_LOGIN_EXEMPT_PREFIXES):
        return None
    if not session.get("hr_username"):
        return redirect(url_for("hr_login", next=request.path))
    if session.get("hr_role") == "approver":
        allowed_prefixes = APPROVER_ALWAYS_ALLOWED_PREFIXES
        for flag, prefixes in APPROVER_CAPABILITY_PREFIXES.items():
            if session.get(flag) == "Y":
                allowed_prefixes += prefixes
        if not request.path.startswith(allowed_prefixes):
            return redirect(_approver_home())
    return None


@app.context_processor
def inject_globals():
    today = datetime.date.today()
    return {"month_names": MONTH_NAMES, "cur_year": today.year, "cur_month": today.month,
            "hr_username": session.get("hr_username"), "hr_role": session.get("hr_role"),
            "can_approve_leave": session.get("can_approve_leave"),
            "can_approve_appraisal": session.get("can_approve_appraisal"),
            "can_approve_ot": session.get("can_approve_ot")}


def employed_this_month(db, year, month, select_cols="emp_id"):
    """Active staff, plus anyone whose last working day falls on/after the
    start of this month - so someone who resigns mid-month still shows up
    for attendance/payroll for the month they actually worked part of.
    Someone with no last_working_day recorded is only included via Active
    status, since there's nothing to say they were still here this month.
    Excludes anyone whose Date Joined is after the end of this month, so a
    new hire doesn't show up on attendance/payroll for months before they
    actually started (mirrors the Staff Portal's own _was_employed check)."""
    first_of_month = f"{year:04d}-{month:02d}-01"
    last_of_month = f"{year:04d}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"
    return db.execute(
        f"""SELECT {select_cols} FROM employees
            WHERE (status='Active'
                   OR (last_working_day IS NOT NULL AND last_working_day != '' AND last_working_day >= ?))
              AND (date_joined IS NULL OR date_joined = '' OR date_joined <= ?)
            ORDER BY emp_id""",
        (first_of_month, last_of_month),
    ).fetchall()


# ---------------- Staff Portal: Auth ----------------

def portal_login_required(view):
    @functools.wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("portal_emp_id"):
            return redirect(url_for("portal_login", next=request.path))
        if request.method == "POST" and session.get("hr_username") and not session.get("portal_self_login"):
            # HR is previewing this employee's portal (portal_preview),
            # not the employee's own real login - block writes so a "just
            # looking" click can't silently overwrite the employee's real
            # data (this is exactly how an HR preview session once saved
            # HR's own email/phone/address onto an employee's record).
            # portal_self_login (set only by portal_login() itself) means
            # this employee's own login legitimately also carries HR access
            # (e.g. Mr Kee), so their own writes are not blocked.
            dest = request.referrer or url_for("portal_dashboard")
            sep = "&" if "?" in dest else "?"
            return redirect(f"{dest}{sep}preview_blocked=1")
        return view(*args, **kwargs)
    return wrapped


@app.route("/hr/go-to-portal")
def hr_go_to_portal():
    """The reverse of the Staff Portal's own merged login: an HR/approver
    user whose account is linked to an employee record (via that
    employee's hr_username, e.g. Linda = L002) jumps straight into their
    own Staff Portal without a second login. Does NOT clear the session
    (hr_username stays), and sets portal_self_login so the write-block
    that protects HR "view as employee" previews doesn't wrongly apply to
    this genuine self-access. Falls back to the normal portal login page
    if this HR account isn't linked to any employee."""
    db = get_db()
    emp = db.execute("SELECT emp_id FROM employees WHERE hr_username=?", (session.get("hr_username"),)).fetchone()
    if emp is None:
        return redirect(url_for("portal_login"))
    session["portal_emp_id"] = emp["emp_id"]
    session["portal_self_login"] = True
    return redirect(url_for("portal_dashboard"))


@app.route("/portal/login", methods=["GET", "POST"])
def portal_login():
    db = get_db()
    error = None
    if request.method == "POST":
        emp_id = request.form.get("emp_id", "").strip().upper()
        password = request.form.get("password", "")
        emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
        if emp is None or not emp["portal_password_hash"]:
            error = "Unknown Employee ID or portal account not yet activated. Contact HR."
        elif not check_password_hash(emp["portal_password_hash"], password):
            error = "Incorrect password."
        else:
            session.clear()
            session["portal_emp_id"] = emp_id
            # If this employee is also linked to an HR/approver account (e.g.
            # Mr Kee = K003 + hr_users 'kee'), one login now grants both -
            # they never need the separate /hr/login. portal_self_login marks
            # this as a genuine self-login (not an HR admin previewing via
            # portal_preview()), so portal_login_required knows not to block
            # their own writes below.
            if emp["hr_username"]:
                hr_user = db.execute("SELECT * FROM hr_users WHERE username=?", (emp["hr_username"],)).fetchone()
                if hr_user:
                    session["hr_username"] = hr_user["username"]
                    session["hr_role"] = hr_user["role"]
                    session["can_approve_leave"] = hr_user["can_approve_leave"]
                    session["can_approve_appraisal"] = hr_user["can_approve_appraisal"]
                    session["can_approve_ot"] = hr_user["can_approve_ot"]
                    session["portal_self_login"] = True
            next_url = request.form.get("next") or url_for("portal_dashboard")
            return redirect(next_url)
    return render_template("portal_login.html", error=error, next=request.args.get("next", ""))


@app.route("/portal/logout")
def portal_logout():
    session.clear()
    return redirect(url_for("portal_login"))


@app.route("/employees/<emp_id>/portal-preview")
def portal_preview(emp_id):
    """Lets HR (role='admin') see exactly what an employee sees on their
    Staff Portal, without knowing/resetting their password. Sets
    portal_emp_id alongside the existing hr_username/hr_role (does NOT
    clear the session like a real portal login would), so the HR session
    is still there to return to via Exit Preview. Naturally unreachable by
    an 'approver' account since this route sits outside their allowed
    prefixes."""
    db = get_db()
    if db.execute("SELECT 1 FROM employees WHERE emp_id=?", (emp_id,)).fetchone() is None:
        return "Employee not found", 404
    session["portal_emp_id"] = emp_id
    return redirect(url_for("portal_dashboard"))


@app.route("/portal/exit-preview")
def portal_exit_preview():
    """Ends an HR preview session (pops portal_emp_id only) without
    touching the HR login, unlike portal_logout which clears everything -
    that stays for real employees signing out of their own account."""
    return_emp_id = session.pop("portal_emp_id", None)
    if session.get("hr_username") and return_emp_id:
        return redirect(url_for("edit_employee", emp_id=return_emp_id))
    return redirect(url_for("portal_login"))


@app.route("/portal/forgot-password", methods=["GET", "POST"])
def portal_forgot_password():
    db = get_db()
    error = None
    success = False
    if request.method == "POST":
        emp_id = request.form.get("emp_id", "").strip().upper()
        ic_input = request.form.get("ic_passport_no", "").strip().replace("-", "").replace(" ", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
        stored_ic = (emp["ic_passport_no"] or "").replace("-", "").replace(" ", "") if emp else ""
        # Identity is verified with Employee ID + IC/Passport No. (both already on
        # file with HR) rather than an email link, since this system has no email
        # server set up - matches how a small company without IT infra can still
        # offer genuine self-service instead of "ask HR every time".
        if emp is None or not ic_input or ic_input.upper() != stored_ic.upper():
            error = "Employee ID and IC/Passport No. don't match our records. Contact HR if you're stuck."
        elif len(new_password) < 4:
            error = "New password must be at least 4 characters."
        elif new_password != confirm_password:
            error = "New password and confirmation don't match."
        else:
            db.execute(
                "UPDATE employees SET portal_password_hash=? WHERE emp_id=?",
                (generate_password_hash(new_password), emp_id),
            )
            db.commit()
            success = True
    return render_template("portal_forgot_password.html", error=error, success=success)


def current_portal_employee(db):
    return db.execute("SELECT * FROM employees WHERE emp_id=?", (session["portal_emp_id"],)).fetchone()


@app.route("/")
def index():
    if session.get("hr_role") == "approver":
        return redirect(_approver_home())
    today = datetime.date.today()
    return redirect(url_for("run_payroll", year=today.year, month=today.month))


# ---------------- HR Login ----------------
# Separate from Staff Portal auth above - hr_users grants access to the
# whole HR/payroll side, employees.portal_password_hash only ever grants
# an employee access to their own portal page.

@app.route("/hr/login", methods=["GET", "POST"])
def hr_login():
    db = get_db()
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        user = db.execute("SELECT * FROM hr_users WHERE username=?", (username,)).fetchone()
        if user is None or not check_password_hash(user["password_hash"], password):
            error = "Incorrect username or password."
        else:
            session.clear()
            session["hr_username"] = user["username"]
            session["hr_role"] = user["role"]
            session["can_approve_leave"] = user["can_approve_leave"]
            session["can_approve_appraisal"] = user["can_approve_appraisal"]
            session["can_approve_ot"] = user["can_approve_ot"]
            next_url = request.form.get("next") or url_for("index")
            # Never redirect off-site or back into the login route itself.
            if not next_url.startswith("/") or next_url.startswith("/hr/login"):
                next_url = url_for("index")
            return redirect(next_url)
    no_hr_users = db.execute("SELECT 1 FROM hr_users LIMIT 1").fetchone() is None
    return render_template("hr_login.html", error=error, next=request.args.get("next", ""),
                            no_hr_users=no_hr_users)


@app.route("/hr/logout")
def hr_logout():
    session.clear()
    return redirect(url_for("hr_login"))


@app.route("/hr/setup", methods=["GET", "POST"])
def hr_setup():
    """First-run only: creates the very first HR (role='admin') account on
    a brand new deployment that has no hr_users yet (e.g. right after
    deploying to Railway/Render, where the database bootstraps empty on
    purpose - no employee/payroll data or credentials ever leave the local
    machine via GitHub). Locks itself the moment any hr_users row exists,
    so it can't be used to create extra unauthorized accounts later."""
    db = get_db()
    if db.execute("SELECT 1 FROM hr_users LIMIT 1").fetchone() is not None:
        return redirect(url_for("hr_login"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        full_name = request.form.get("full_name", "").strip()
        if not username or not password or not full_name:
            error = "Username, password, and full name are all required."
        elif len(password) < 4:
            error = "Password must be at least 4 characters."
        else:
            db.execute(
                "INSERT INTO hr_users (username, password_hash, full_name, created_at, role) VALUES (?,?,?,?,'admin')",
                (username, generate_password_hash(password), full_name,
                 datetime.datetime.now().isoformat(timespec="seconds")),
            )
            db.commit()
            session.clear()
            session["hr_username"] = username
            session["hr_role"] = "admin"
            return redirect(url_for("index"))
    return render_template("hr_setup.html", error=error)


@app.route("/hr/change-password", methods=["GET", "POST"])
def hr_change_password():
    db = get_db()
    error = None
    success = False
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        user = db.execute("SELECT * FROM hr_users WHERE username=?", (session["hr_username"],)).fetchone()
        if not check_password_hash(user["password_hash"], current_password):
            error = "Current password is incorrect."
        elif len(new_password) < 4:
            error = "New password must be at least 4 characters."
        elif new_password != confirm_password:
            error = "New password and confirmation don't match."
        else:
            db.execute("UPDATE hr_users SET password_hash=? WHERE username=?",
                       (generate_password_hash(new_password), session["hr_username"]))
            db.commit()
            success = True
    return render_template("hr_change_password.html", error=error, success=success)


# ---------------- Employees ----------------

TEXT_FIELDS = ["full_name", "ic_passport_no", "date_of_birth", "marital_status",
               "race", "religion", "holiday_state", "epf_no", "socso_no", "tax_no", "skbbk_flag", "eis_flag",
               "department", "position", "status", "work_pattern", "bank_name", "bank_account_no",
               "phone_number", "hp_no", "email", "address",
               "emergency_contact_1_name", "emergency_contact_1_phone", "emergency_contact_1_relationship",
               "emergency_contact_2_name", "emergency_contact_2_phone", "emergency_contact_2_relationship",
               "date_joined", "last_working_day", "probation_end_date",
               "passport_expiry", "work_permit_expiry", "termination_notice_period",
               "confirmation_date", "resignation_date", "appraisal_supervisor_username",
               "leave_approver_username", "hr_username", "ot_approval_required"]
NUM_FIELDS = ["basic_salary", "working_days_week", "working_hours_day",
              "additional_epf_employee", "annual_leave_entitlement", "mc_entitlement",
              "hospitalisation_leave_entitlement", "medical_claim_limit"]
NULLABLE_NUM_FIELDS = ["confirmed_new_salary"]
ALLOWANCE_FIELDS = [
    ("transport_allowance", "transport_allowance_flag", "transport_allowance_effective_date"),
    ("meal_allowance_rate", "meal_allowance_flag", "meal_allowance_effective_date"),
    ("position_allowance", "position_allowance_flag", "position_allowance_effective_date"),
    ("cewi_rate", "cewi_flag", "cewi_effective_date"),
    ("training_incentive", "training_incentive_flag", "training_incentive_effective_date"),
    ("oversea_incentive", "oversea_incentive_flag", "oversea_incentive_effective_date"),
]


def _employee_fields_from_form(f):
    """Builds a {column: value} dict from the employee_edit.html form,
    shared by both the add and edit routes so the field list only lives
    in one place."""
    fields = {}
    for name in TEXT_FIELDS:
        fields[name] = f.get(name) or None
    # Store IC/passport numbers without dashes/spaces so every record is
    # consistent regardless of how it was typed (e.g. "820613-01-5226" and
    # "820613015226" both save as the latter).
    if fields.get("ic_passport_no"):
        fields["ic_passport_no"] = fields["ic_passport_no"].replace("-", "").replace(" ", "")
    for name in NUM_FIELDS:
        fields[name] = float(f.get(name) or 0)
    for name in NULLABLE_NUM_FIELDS:
        raw = f.get(name)
        fields[name] = float(raw) if raw else None
    for amount_field, flag_field, date_field in ALLOWANCE_FIELDS:
        fields[amount_field] = float(f.get(amount_field) or 0)
        fields[flag_field] = f.get(flag_field, "N")
        fields[date_field] = f.get(date_field) or None

    # Confirming an employee with a new salary immediately becomes their
    # actual Basic Salary for payroll - no separate manual step, and no
    # mid-month blending (matches how the source workbook's Salary
    # Adjustment Log always worked: apply the new figure, whole month).
    if fields.get("confirmed_new_salary") is not None:
        fields["basic_salary"] = fields["confirmed_new_salary"]

    return fields


def _eis_applies(dob_str, eis_flag):
    """Same 18-59 working-age rule as calc_eis() in payroll_calc.py, plus the
    manual eis_flag override, just based on today's date rather than a
    specific payroll month - this is only for the at-a-glance Employee
    Master column, not for actual payroll."""
    if eis_flag == "N":
        return False
    if not dob_str:
        return None
    dob = datetime.date.fromisoformat(dob_str)
    age = (datetime.date.today() - dob).days // 365
    return 18 <= age < 60


@app.route("/employees")
def employees():
    db = get_db()
    view = request.args.get("view", "active")
    if view == "inactive":
        rows = db.execute(
            "SELECT * FROM employees WHERE status='Inactive' ORDER BY emp_id"
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM employees WHERE status != 'Inactive' ORDER BY emp_id"
        ).fetchall()
    inactive_count = db.execute(
        "SELECT COUNT(*) AS c FROM employees WHERE status='Inactive'"
    ).fetchone()["c"]
    eis_applies = {e["emp_id"]: _eis_applies(e["date_of_birth"], e["eis_flag"]) for e in rows}
    depts = sorted({e["department"] for e in rows if e["department"]})
    positions = sorted({e["position"] for e in rows if e["position"]})
    return render_template("employees.html", employees=rows, view=view,
                            inactive_count=inactive_count, eis_applies=eis_applies,
                            depts=depts, positions=positions)


@app.route("/employees/new", methods=["GET", "POST"])
def add_employee():
    db = get_db()
    error = None
    if request.method == "POST":
        emp_id = (request.form.get("emp_id") or "").strip()
        if not emp_id:
            error = "Employee ID is required."
        elif db.execute("SELECT 1 FROM employees WHERE emp_id=?", (emp_id,)).fetchone():
            error = f"Employee ID '{emp_id}' already exists."
        else:
            fields = _employee_fields_from_form(request.form)
            columns = ["emp_id"] + list(fields.keys())
            placeholders = ",".join(["?"] * len(columns))
            db.execute(
                f"INSERT INTO employees ({','.join(columns)}) VALUES ({placeholders})",
                [emp_id] + list(fields.values()),
            )
            # Seed Salary History with an "Initial" record tied to Date Joined,
            # so every employee's history starts from a known baseline rather
            # than only showing changes made after this feature existed.
            starting_salary = fields.get("basic_salary") or 0
            db.execute(
                """INSERT INTO salary_history (emp_id, effective_date, old_salary, new_salary, increment, reason, recorded_at)
                   VALUES (?,?,?,?,?,?,?)""",
                (emp_id, fields.get("date_joined") or datetime.date.today().isoformat(),
                 starting_salary, starting_salary, 0, "Initial",
                 datetime.datetime.now().isoformat(timespec="seconds")),
            )
            # Same idea for Probation Extension - if a Probation End Date was
            # set at creation, log it as the baseline so the history isn't
            # blank until the first actual extension.
            if fields.get("probation_end_date"):
                db.execute(
                    """INSERT INTO probation_extensions (emp_id, previous_end_date, new_end_date, reason, extended_at)
                       VALUES (?,?,?,?,?)""",
                    (emp_id, "(not set)", fields["probation_end_date"], "Initial",
                     datetime.datetime.now().isoformat(timespec="seconds")),
                )
            db.commit()
            return redirect(url_for("edit_employee", emp_id=emp_id))
    return render_template("employee_edit.html", emp={}, is_new=True, error=error,
                            race_options=RACE_OPTIONS, religion_options=RELIGION_OPTIONS,
                            holiday_state_options=HOLIDAY_STATE_OPTIONS)


@app.route("/employees/<emp_id>/edit", methods=["GET", "POST"])
def edit_employee(emp_id):
    db = get_db()
    if request.method == "POST":
        fields = _employee_fields_from_form(request.form)
        new_password = request.form.get("portal_password", "").strip()
        if new_password:
            fields["portal_password_hash"] = generate_password_hash(new_password)
        # If Probation End Date is being set for the first time (was empty,
        # now isn't), log it as the Initial baseline in Probation Extension
        # history - same "normal setup now has a starting record" logic as
        # the Salary History Initial row.
        if fields.get("probation_end_date"):
            was_empty = db.execute(
                "SELECT 1 FROM employees WHERE emp_id=? AND (probation_end_date IS NULL OR probation_end_date='')",
                (emp_id,),
            ).fetchone()
            if was_empty:
                db.execute(
                    """INSERT INTO probation_extensions (emp_id, previous_end_date, new_end_date, reason, extended_at)
                       VALUES (?,?,?,?,?)""",
                    (emp_id, "(not set)", fields["probation_end_date"], "Initial",
                     datetime.datetime.now().isoformat(timespec="seconds")),
                )
        set_clause = ",".join(f"{col}=?" for col in fields)
        db.execute(
            f"UPDATE employees SET {set_clause} WHERE emp_id=?",
            list(fields.values()) + [emp_id],
        )
        db.commit()
        return redirect(url_for("edit_employee", emp_id=emp_id))

    emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    tax_profile = db.execute("SELECT * FROM tax_profile WHERE emp_id=?", (emp_id,)).fetchone()
    extensions = db.execute(
        "SELECT * FROM probation_extensions WHERE emp_id=? ORDER BY extended_at DESC",
        (emp_id,),
    ).fetchall()
    salary_history = db.execute(
        """SELECT * FROM salary_history WHERE emp_id=?
           ORDER BY (reason = 'Initial') DESC, effective_date DESC, recorded_at DESC""",
        (emp_id,),
    ).fetchall()
    documents = db.execute(
        "SELECT * FROM employee_documents WHERE emp_id=? ORDER BY uploaded_at DESC",
        (emp_id,),
    ).fetchall()

    eis_applies = _eis_applies(emp["date_of_birth"], emp["eis_flag"])

    al_year = int(emp["last_working_day"][:4]) if emp["last_working_day"] else datetime.date.today().year
    prorated_al, al_note = _prorated_al_note(emp, al_year)
    al_used = db.execute(
        "SELECT COALESCE(SUM(al_days),0) AS al FROM attendance_monthly WHERE emp_id=? AND year=?",
        (emp_id, al_year),
    ).fetchone()["al"]
    al_entitlement_effective = prorated_al
    al_balance = (al_entitlement_effective - al_used) if al_entitlement_effective is not None else None

    appraisal_supervisors = db.execute(
        "SELECT username, full_name FROM hr_users WHERE can_approve_appraisal='Y' ORDER BY full_name"
    ).fetchall()
    leave_approvers = db.execute(
        "SELECT username, full_name FROM hr_users WHERE can_approve_leave='Y' ORDER BY full_name"
    ).fetchall()
    hr_accounts = db.execute("SELECT username, full_name FROM hr_users ORDER BY full_name").fetchall()

    return render_template("employee_edit.html", emp=emp, is_new=False, extensions=extensions,
                            salary_history=salary_history, eis_applies=eis_applies,
                            documents=documents, document_types=DOCUMENT_TYPES,
                            al_note=al_note, al_year=al_year, al_entitlement_effective=al_entitlement_effective,
                            al_used=al_used, al_balance=al_balance,
                            race_options=RACE_OPTIONS, religion_options=RELIGION_OPTIONS,
                            holiday_state_options=HOLIDAY_STATE_OPTIONS,
                            appraisal_supervisors=appraisal_supervisors, leave_approvers=leave_approvers,
                            hr_accounts=hr_accounts,
                            tax_profile=tax_profile)


@app.route("/employees/<emp_id>/tax-profile/update", methods=["POST"])
def update_tax_profile(emp_id):
    db = get_db()
    emp = db.execute("SELECT emp_id FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None:
        return "Employee not found", 404
    f = request.form
    tax_category = f.get("tax_category") or "Single"
    children_full = int(f.get("children_full_relief") or 0)
    children_half = int(f.get("children_half_relief") or 0)
    tp1_submitted = "Y" if f.get("tp1_submitted") else ""
    tp1_date = f.get("tp1_date") or None
    zakat_paid_ytd = float(f.get("zakat_paid_ytd") or 0)
    tp3_submitted = "Y" if f.get("tp3_submitted") else ""
    tp3_date = f.get("tp3_date") or None
    tp3_prior_gross = float(f.get("tp3_prior_gross") or 0)
    tp3_prior_epf_employee = float(f.get("tp3_prior_epf_employee") or 0)
    tp3_prior_pcb = float(f.get("tp3_prior_pcb") or 0)
    db.execute(
        """INSERT INTO tax_profile (
               emp_id, tax_category, children_full_relief, children_half_relief,
               tp1_submitted, tp1_date, zakat_paid_ytd,
               tp3_submitted, tp3_date, tp3_prior_gross, tp3_prior_epf_employee, tp3_prior_pcb
           ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
           ON CONFLICT(emp_id) DO UPDATE SET
             tax_category=excluded.tax_category,
             children_full_relief=excluded.children_full_relief,
             children_half_relief=excluded.children_half_relief,
             tp1_submitted=excluded.tp1_submitted, tp1_date=excluded.tp1_date,
             zakat_paid_ytd=excluded.zakat_paid_ytd,
             tp3_submitted=excluded.tp3_submitted, tp3_date=excluded.tp3_date,
             tp3_prior_gross=excluded.tp3_prior_gross,
             tp3_prior_epf_employee=excluded.tp3_prior_epf_employee,
             tp3_prior_pcb=excluded.tp3_prior_pcb""",
        (emp_id, tax_category, children_full, children_half, tp1_submitted, tp1_date, zakat_paid_ytd,
         tp3_submitted, tp3_date, tp3_prior_gross, tp3_prior_epf_employee, tp3_prior_pcb),
    )
    db.commit()
    return redirect(url_for("edit_employee", emp_id=emp_id))


@app.route("/employees/<emp_id>/extend-probation", methods=["POST"])
def extend_probation(emp_id):
    db = get_db()
    emp = db.execute("SELECT probation_end_date FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None:
        return "Employee not found", 404
    new_end_date = request.form.get("new_end_date")
    reason = request.form.get("reason") or None
    if not new_end_date:
        return redirect(url_for("edit_employee", emp_id=emp_id))
    db.execute(
        """INSERT INTO probation_extensions (emp_id, previous_end_date, new_end_date, reason, extended_at)
           VALUES (?,?,?,?,?)""",
        (emp_id, emp["probation_end_date"] or "(not set)", new_end_date, reason,
         datetime.datetime.now().isoformat(timespec="seconds")),
    )
    db.execute("UPDATE employees SET probation_end_date=? WHERE emp_id=?", (new_end_date, emp_id))
    db.commit()
    return redirect(url_for("edit_employee", emp_id=emp_id))


@app.route("/employees/<emp_id>/probation-extensions/<int:ext_id>/delete", methods=["POST"])
def delete_probation_extension(emp_id, ext_id):
    db = get_db()
    db.execute("DELETE FROM probation_extensions WHERE id=? AND emp_id=?", (ext_id, emp_id))
    db.commit()
    return redirect(url_for("edit_employee", emp_id=emp_id))


@app.route("/employees/<emp_id>/update-salary", methods=["POST"])
def update_salary(emp_id):
    db = get_db()
    emp = db.execute("SELECT basic_salary FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None:
        return "Employee not found", 404
    effective_date = request.form.get("effective_date")
    reason = request.form.get("reason") or None
    new_salary_raw = request.form.get("new_salary")
    if not effective_date or not reason or not new_salary_raw:
        return redirect(url_for("edit_employee", emp_id=emp_id))
    old_salary = emp["basic_salary"]
    new_salary = float(new_salary_raw)
    db.execute(
        """INSERT INTO salary_history (emp_id, effective_date, old_salary, new_salary, increment, reason, recorded_at)
           VALUES (?,?,?,?,?,?,?)""",
        (emp_id, effective_date, old_salary, new_salary, new_salary - old_salary, reason,
         datetime.datetime.now().isoformat(timespec="seconds")),
    )
    db.execute("UPDATE employees SET basic_salary=? WHERE emp_id=?", (new_salary, emp_id))
    if reason == "Confirmation":
        # A salary update logged with Reason=Confirmation also settles the
        # Confirmation Date/Salary fields the Confirmation Letter reads from -
        # one action instead of filling in the same numbers twice.
        db.execute(
            "UPDATE employees SET confirmation_date=?, confirmed_new_salary=? WHERE emp_id=?",
            (effective_date, new_salary, emp_id),
        )
    db.commit()
    return redirect(url_for("edit_employee", emp_id=emp_id))


def _resync_basic_salary_from_history(db, emp_id):
    """Keeps employees.basic_salary matching whichever salary_history row is
    most recent (by recorded_at), after that history log is corrected or a
    row is removed - so editing/deleting a record doesn't leave Basic
    Salary pointing at a value that's no longer in the log."""
    latest = db.execute(
        "SELECT new_salary FROM salary_history WHERE emp_id=? ORDER BY recorded_at DESC LIMIT 1",
        (emp_id,),
    ).fetchone()
    if latest is not None:
        db.execute("UPDATE employees SET basic_salary=? WHERE emp_id=?", (latest["new_salary"], emp_id))


@app.route("/employees/<emp_id>/salary-history/<int:hist_id>/update", methods=["POST"])
def update_salary_history(emp_id, hist_id):
    db = get_db()
    row = db.execute(
        "SELECT * FROM salary_history WHERE id=? AND emp_id=?", (hist_id, emp_id)
    ).fetchone()
    if row is None:
        return "Salary history record not found", 404
    effective_date = request.form.get("effective_date")
    reason = request.form.get("reason") or None
    new_salary_raw = request.form.get("new_salary")
    old_salary_raw = request.form.get("old_salary")
    if not effective_date or not reason or not new_salary_raw or not old_salary_raw:
        return redirect(url_for("edit_employee", emp_id=emp_id))
    new_salary = float(new_salary_raw)
    old_salary = float(old_salary_raw)
    db.execute(
        """UPDATE salary_history SET effective_date=?, old_salary=?, new_salary=?, increment=?, reason=?
           WHERE id=?""",
        (effective_date, old_salary, new_salary, new_salary - old_salary, reason, hist_id),
    )
    _resync_basic_salary_from_history(db, emp_id)
    db.commit()
    return redirect(url_for("edit_employee", emp_id=emp_id))


@app.route("/employees/<emp_id>/salary-history/<int:hist_id>/delete", methods=["POST"])
def delete_salary_history(emp_id, hist_id):
    db = get_db()
    db.execute("DELETE FROM salary_history WHERE id=? AND emp_id=?", (hist_id, emp_id))
    _resync_basic_salary_from_history(db, emp_id)
    db.commit()
    return redirect(url_for("edit_employee", emp_id=emp_id))


# ---------------- Employee Photo ----------------

@app.route("/employees/<emp_id>/photo/upload", methods=["POST"])
def upload_photo(emp_id):
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None:
        return "Employee not found", 404

    file = request.files.get("photo")
    if file is None or file.filename == "":
        return redirect(url_for("edit_employee", emp_id=emp_id))

    original_name = secure_filename(file.filename)
    ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""
    if ext not in ALLOWED_PHOTO_EXTENSIONS:
        return "File type not allowed. Use JPG or PNG.", 400

    emp_dir = os.path.join(UPLOAD_DIR, emp_id)
    os.makedirs(emp_dir, exist_ok=True)

    # Replace any previous photo so we don't accumulate old ones.
    if emp["photo_path"]:
        old_path = os.path.join(emp_dir, emp["photo_path"])
        if os.path.exists(old_path):
            os.remove(old_path)

    stored_name = f"photo_{uuid.uuid4().hex}.{ext}"
    file.save(os.path.join(emp_dir, stored_name))
    db.execute("UPDATE employees SET photo_path=? WHERE emp_id=?", (stored_name, emp_id))
    db.commit()
    return redirect(url_for("edit_employee", emp_id=emp_id))


@app.route("/employees/<emp_id>/photo")
def employee_photo(emp_id):
    db = get_db()
    emp = db.execute("SELECT photo_path FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None or not emp["photo_path"]:
        abort(404)
    return send_from_directory(os.path.join(UPLOAD_DIR, emp_id), emp["photo_path"])


# ---------------- Employee Documents ----------------
# Letter of Employment, Confirmation Letter, Resignation Letter, e-Stamping
# Certificate, etc. - uploaded file kept per employee. Files live under
# uploads/<emp_id>/<uuid>_<original filename>; the DB row just indexes them.

@app.route("/employees/<emp_id>/documents/upload", methods=["POST"])
def upload_document(emp_id):
    db = get_db()
    if db.execute("SELECT 1 FROM employees WHERE emp_id=?", (emp_id,)).fetchone() is None:
        return "Employee not found", 404

    doc_type = request.form.get("doc_type") or "Other"
    notes = request.form.get("notes") or None
    file = request.files.get("file")
    if file is None or file.filename == "":
        return redirect(url_for("edit_employee", emp_id=emp_id))

    original_name = secure_filename(file.filename)
    ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""
    if ext not in ALLOWED_DOC_EXTENSIONS:
        return "File type not allowed. Use PDF, Word, or an image (JPG/PNG).", 400

    emp_dir = os.path.join(UPLOAD_DIR, emp_id)
    os.makedirs(emp_dir, exist_ok=True)
    stored_name = f"{uuid.uuid4().hex}_{original_name}"
    file.save(os.path.join(emp_dir, stored_name))

    db.execute(
        """INSERT INTO employee_documents (emp_id, doc_type, original_name, stored_name, notes, uploaded_at)
           VALUES (?,?,?,?,?,?)""",
        (emp_id, doc_type, original_name, stored_name, notes,
         datetime.datetime.now().isoformat(timespec="seconds")),
    )
    db.commit()
    return redirect(url_for("edit_employee", emp_id=emp_id))


@app.route("/employees/<emp_id>/documents/<int:doc_id>/download")
def download_document(emp_id, doc_id):
    db = get_db()
    doc = db.execute(
        "SELECT * FROM employee_documents WHERE id=? AND emp_id=?", (doc_id, emp_id)
    ).fetchone()
    if doc is None:
        abort(404)
    return send_from_directory(
        os.path.join(UPLOAD_DIR, emp_id), doc["stored_name"],
        as_attachment=True, download_name=doc["original_name"],
    )


@app.route("/employees/<emp_id>/documents/<int:doc_id>/view")
def view_document(emp_id, doc_id):
    # Same file as download, but without as_attachment so the browser renders
    # it inline (PDFs and images) instead of forcing a Save As dialog. File
    # types the browser can't display natively (e.g. .docx) still fall back
    # to a download - that's the browser's own behavior, not ours to change.
    db = get_db()
    doc = db.execute(
        "SELECT * FROM employee_documents WHERE id=? AND emp_id=?", (doc_id, emp_id)
    ).fetchone()
    if doc is None:
        abort(404)
    return send_from_directory(
        os.path.join(UPLOAD_DIR, emp_id), doc["stored_name"],
        as_attachment=False, download_name=doc["original_name"],
    )


@app.route("/employees/<emp_id>/documents/<int:doc_id>/delete", methods=["POST"])
def delete_document(emp_id, doc_id):
    db = get_db()
    doc = db.execute(
        "SELECT * FROM employee_documents WHERE id=? AND emp_id=?", (doc_id, emp_id)
    ).fetchone()
    if doc is not None:
        file_path = os.path.join(UPLOAD_DIR, emp_id, doc["stored_name"])
        if os.path.exists(file_path):
            os.remove(file_path)
        db.execute("DELETE FROM employee_documents WHERE id=?", (doc_id,))
        db.commit()
    return redirect(url_for("edit_employee", emp_id=emp_id))


# ---------------- Attendance ----------------

# Each pattern maps weekday (Monday=0 ... Sunday=6) to how much of a
# working day it counts as. A weekday missing from the map is a day off
# (0). "5.5-day" credits Saturday as a half-day, matching a Mon-Fri full
# day + half-day-every-Saturday schedule (a fixed, calendar-predictable
# rule - unlike an *alternating* Saturday-off schedule, which depends on
# an arrangement the calendar can't know and must stay 'Manual').
WORK_PATTERN_WEEKDAY_WEIGHTS = {
    "5-day (Mon-Fri)": {0: 1, 1: 1, 2: 1, 3: 1, 4: 1},
    "5.5-day (Mon-Fri + half-day Sat)": {0: 1, 1: 1, 2: 1, 3: 1, 4: 1, 5: 0.5},
    "6-day (Mon-Sat)": {0: 1, 1: 1, 2: 1, 3: 1, 4: 1, 5: 1},
}


def _calc_working_days_from_pattern(db, year, month, work_pattern):
    """Deterministic Working Days in Month and PH count for an employee on a
    fixed weekly schedule - pure calendar math, no guessing - using the
    Public Holidays list (every row applies to everyone; "Remarks" is just a
    free-text note, not a state filter) as the source of truth. Returns None
    for 'Manual' (or anything not recognized), so the caller falls back to
    the normal blank/manual entry. Alternating/irregular Saturday schedules
    must stay 'Manual': confirmed against a real swipe-card export
    (Shamsury/S002) that they can't be predicted from a calendar alone."""
    weekday_weights = WORK_PATTERN_WEEKDAY_WEIGHTS.get(work_pattern)
    if weekday_weights is None:
        return None

    days_in_month = calendar.monthrange(year, month)[1]
    workdays = {
        datetime.date(year, month, day): weekday_weights[datetime.date(year, month, day).weekday()]
        for day in range(1, days_in_month + 1)
        if datetime.date(year, month, day).weekday() in weekday_weights
    }

    holiday_rows = db.execute(
        "SELECT date FROM public_holidays WHERE date LIKE ?",
        (f"{year:04d}-{month:02d}-%",),
    ).fetchall()
    ph_days = 0
    for h in holiday_rows:
        try:
            h_date = datetime.date.fromisoformat(h["date"])
        except ValueError:
            continue
        if h_date in workdays:
            ph_days += workdays.pop(h_date)

    return {"working_days": sum(workdays.values()), "ph_days": ph_days}


@app.route("/attendance/<int:year>/<int:month>", methods=["GET", "POST"])
def attendance(year, month):
    db = get_db()
    if request.method == "POST":
        emp_ids = [r["emp_id"] for r in employed_this_month(db, year, month)]
        fields = ["days_worked", "al_days", "mc_days", "hl_days", "ul_days",
                  "other_paid_leave", "ph_days", "off_days", "rest_days",
                  "absent_days", "working_days_in_month",
                  "ot_hours_1_5", "ot_hours_2_0", "ot_hours_3_0",
                  "meal_eligible_days"]
        for emp_id in emp_ids:
            values = [float(request.form.get(f"{f}__{emp_id}", 0) or 0) for f in fields]
            db.execute(
                f"""INSERT INTO attendance_monthly (emp_id, year, month, {", ".join(fields)})
                    VALUES (?,?,?,{",".join(["?"] * len(fields))})
                    ON CONFLICT(emp_id, year, month) DO UPDATE SET
                    {", ".join(f"{f}=excluded.{f}" for f in fields)}""",
                [emp_id, year, month] + values,
            )
            var_amt = float(request.form.get(f"variable_allowance__{emp_id}", 0) or 0)
            var_flag = request.form.get(f"variable_allowance_flag__{emp_id}", "N")
            other_ded = float(request.form.get(f"other_deduction__{emp_id}", 0) or 0)
            other_ded_desc = request.form.get(f"other_deduction_desc__{emp_id}", "") or None
            db.execute(
                """INSERT INTO monthly_adjustments (
                       emp_id, year, month, variable_allowance, variable_allowance_flag,
                       other_deduction, other_deduction_desc)
                   VALUES (?,?,?,?,?,?,?)
                   ON CONFLICT(emp_id, year, month) DO UPDATE SET
                     variable_allowance=excluded.variable_allowance,
                     variable_allowance_flag=excluded.variable_allowance_flag,
                     other_deduction=excluded.other_deduction,
                     other_deduction_desc=excluded.other_deduction_desc""",
                (emp_id, year, month, var_amt, var_flag, other_ded, other_ded_desc),
            )
        db.commit()
        return redirect(url_for("attendance", year=year, month=month))

    emps = employed_this_month(
        db, year, month,
        "emp_id, full_name, meal_allowance_flag, cewi_flag, work_pattern",
    )
    att_rows = {
        r["emp_id"]: r for r in db.execute(
            "SELECT * FROM attendance_monthly WHERE year=? AND month=?", (year, month)
        ).fetchall()
    }
    adj_rows = {
        r["emp_id"]: r for r in db.execute(
            "SELECT * FROM monthly_adjustments WHERE year=? AND month=?", (year, month)
        ).fetchall()
    }
    # For anyone on a fixed Work Pattern who doesn't have this month's
    # attendance saved yet, suggest Working Days in Month and PH days
    # automatically instead of leaving them blank - still just a pre-fill,
    # editable/overridable before saving like any other value on this page.
    suggested_working_days = {}
    suggested_ph_days = {}
    for e in emps:
        if e["emp_id"] in att_rows:
            continue
        suggestion = _calc_working_days_from_pattern(
            db, year, month, e["work_pattern"]
        )
        if suggestion is not None:
            suggested_working_days[e["emp_id"]] = suggestion["working_days"]
            suggested_ph_days[e["emp_id"]] = suggestion["ph_days"]
    return render_template("attendance.html", employees=emps, att=att_rows, adj=adj_rows,
                            year=year, month=month, suggested_working_days=suggested_working_days,
                            suggested_ph_days=suggested_ph_days)


DAY_TYPES = ["WORKED", "OFF", "REST", "PH", "AL", "MC", "HL", "UL", "OTHER_PAID"]


def _sync_daily_to_monthly(db, emp_id, year, month):
    """Recomputes attendance_monthly's aggregate columns from this
    employee/month's attendance_daily rows and upserts them - the same
    row payroll already reads, so a daily-entered month feeds payroll
    exactly like a directly-typed monthly total would."""
    rows = db.execute(
        "SELECT * FROM attendance_daily WHERE emp_id=? AND date LIKE ?",
        (emp_id, f"{year:04d}-{month:02d}-%"),
    ).fetchall()
    counts = {t: 0 for t in DAY_TYPES}
    for r in rows:
        counts[r["day_type"]] = counts.get(r["day_type"], 0) + 1
    ot_1_5 = sum(r["ot_hours_1_5"] or 0 for r in rows)
    ot_2_0 = sum(r["ot_hours_2_0"] or 0 for r in rows)
    ot_3_0 = sum(r["ot_hours_3_0"] or 0 for r in rows)
    meal_days = sum(1 for r in rows if r["meal_allowance_flag"] == "Y")
    working_days_in_month = counts["WORKED"] + counts["AL"] + counts["MC"] + counts["HL"] + counts["OTHER_PAID"] + counts["PH"]

    db.execute(
        """INSERT INTO attendance_monthly (
               emp_id, year, month, days_worked, al_days, mc_days, hl_days, ul_days,
               other_paid_leave, ph_days, off_days, rest_days, working_days_in_month,
               ot_hours_1_5, ot_hours_2_0, ot_hours_3_0, meal_eligible_days
           ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
           ON CONFLICT(emp_id, year, month) DO UPDATE SET
             days_worked=excluded.days_worked, al_days=excluded.al_days, mc_days=excluded.mc_days,
             hl_days=excluded.hl_days, ul_days=excluded.ul_days, other_paid_leave=excluded.other_paid_leave,
             ph_days=excluded.ph_days, off_days=excluded.off_days, rest_days=excluded.rest_days,
             working_days_in_month=excluded.working_days_in_month,
             ot_hours_1_5=excluded.ot_hours_1_5, ot_hours_2_0=excluded.ot_hours_2_0,
             ot_hours_3_0=excluded.ot_hours_3_0, meal_eligible_days=excluded.meal_eligible_days""",
        (emp_id, year, month, counts["WORKED"], counts["AL"], counts["MC"], counts["HL"], counts["UL"],
         counts["OTHER_PAID"], counts["PH"], counts["OFF"], counts["REST"], working_days_in_month,
         ot_1_5, ot_2_0, ot_3_0, meal_days),
    )


@app.route("/attendance-daily/<emp_id>/<int:year>/<int:month>", methods=["GET", "POST"])
def attendance_daily(emp_id, year, month):
    """Day-by-day attendance entry for one employee - HR fills in each
    day's Time In/Out, meal allowance, and OT hours, matching the layout
    of a real daily attendance sheet, instead of typing straight-to-
    monthly totals. Saving recomputes this month's attendance_monthly row
    from the daily entries (see _sync_daily_to_monthly), so it feeds
    payroll exactly like the bulk Attendance page would."""
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None:
        return "Employee not found", 404
    days_in_month = calendar.monthrange(year, month)[1]

    if request.method == "POST":
        for day in range(1, days_in_month + 1):
            date_str = f"{year:04d}-{month:02d}-{day:02d}"
            day_type = request.form.get(f"day_type__{day}") or "WORKED"
            if day_type not in DAY_TYPES:
                day_type = "WORKED"
            time_in = request.form.get(f"time_in__{day}") or None
            time_out = request.form.get(f"time_out__{day}") or None
            meal_flag = "Y" if request.form.get(f"meal__{day}") else "N"
            ot_1_5 = float(request.form.get(f"ot_1_5__{day}", 0) or 0)
            ot_2_0 = float(request.form.get(f"ot_2_0__{day}", 0) or 0)
            ot_3_0 = float(request.form.get(f"ot_3_0__{day}", 0) or 0)
            ot_reason = request.form.get(f"ot_reason__{day}") or None
            db.execute(
                """INSERT INTO attendance_daily (
                       emp_id, date, day_type, time_in, time_out, meal_allowance_flag,
                       ot_hours_1_5, ot_hours_2_0, ot_hours_3_0, ot_reason
                   ) VALUES (?,?,?,?,?,?,?,?,?,?)
                   ON CONFLICT(emp_id, date) DO UPDATE SET
                     day_type=excluded.day_type, time_in=excluded.time_in, time_out=excluded.time_out,
                     meal_allowance_flag=excluded.meal_allowance_flag,
                     ot_hours_1_5=excluded.ot_hours_1_5, ot_hours_2_0=excluded.ot_hours_2_0,
                     ot_hours_3_0=excluded.ot_hours_3_0, ot_reason=excluded.ot_reason""",
                (emp_id, date_str, day_type, time_in, time_out, meal_flag,
                 ot_1_5, ot_2_0, ot_3_0, ot_reason),
            )
        _sync_daily_to_monthly(db, emp_id, year, month)
        db.commit()
        return redirect(url_for("attendance_daily", emp_id=emp_id, year=year, month=month))

    saved = {
        r["date"]: r for r in db.execute(
            "SELECT * FROM attendance_daily WHERE emp_id=? AND date LIKE ? ORDER BY date",
            (emp_id, f"{year:04d}-{month:02d}-%"),
        ).fetchall()
    }
    days = []
    for day in range(1, days_in_month + 1):
        date_obj = datetime.date(year, month, day)
        days.append({
            "day": day, "date": date_obj.isoformat(), "weekday": date_obj.strftime("%a"),
            "row": saved.get(date_obj.isoformat()),
        })
    monthly = db.execute(
        "SELECT * FROM attendance_monthly WHERE emp_id=? AND year=? AND month=?",
        (emp_id, year, month),
    ).fetchone()

    return render_template("attendance_daily.html", emp=emp, year=year, month=month,
                            days=days, day_types=DAY_TYPES, monthly=monthly)


# ---------------- Payroll History ----------------

@app.route("/history")
def history():
    db = get_db()
    rows = db.execute(
        """SELECT year, month, COUNT(*) AS employees,
                  SUM(gross_pay) AS total_gross, SUM(net_pay) AS total_net,
                  MAX(finalized_at) AS finalized_at
           FROM payroll_runs GROUP BY year, month ORDER BY year ASC, month ASC"""
    ).fetchall()
    return render_template("history.html", rows=rows)


# ---------------- Alerts (Confirmation / Passport / Work Permit) ----------------

def _rows_with_days_left(db, date_column, extra_cols="", exclude_if_confirmed=False):
    today = datetime.date.today()
    confirmed_clause = (
        "AND (confirmation_date IS NULL OR confirmation_date = '') " if exclude_if_confirmed else ""
    )
    out = []
    for r in db.execute(
        f"""SELECT emp_id, full_name, department, position {extra_cols}, {date_column} AS the_date
            FROM employees WHERE {date_column} IS NOT NULL AND {date_column} != ''
              AND (status IS NULL OR status != 'Inactive')
              {confirmed_clause}
            ORDER BY {date_column}"""
    ).fetchall():
        end_date = datetime.date.fromisoformat(r["the_date"])
        days_left = (end_date - today).days
        out.append({**dict(r), "days_left": days_left, "overdue": days_left < 0})
    return out


@app.route("/alerts")
def confirmation_due():
    db = get_db()
    today = datetime.date.today()

    due = _rows_with_days_left(db, "probation_end_date", ", date_joined, appraisal_supervisor_username",
                                exclude_if_confirmed=True)
    supervisor_names = {
        r["username"]: r["full_name"]
        for r in db.execute("SELECT username, full_name FROM hr_users").fetchall()
    }
    for r in due:
        r["reports_to"] = supervisor_names.get(r["appraisal_supervisor_username"]) or r["appraisal_supervisor_username"] or "-"
    passport_alerts = _rows_with_days_left(db, "passport_expiry")
    work_permit_alerts = _rows_with_days_left(db, "work_permit_expiry")

    six_months_ago = (today.replace(day=1) - datetime.timedelta(days=183)).isoformat()
    missing = db.execute(
        """SELECT emp_id, full_name, department, position, date_joined
           FROM employees
           WHERE (probation_end_date IS NULL OR probation_end_date = '')
             AND date_joined IS NOT NULL AND date_joined != '' AND date_joined >= ?
             AND (status IS NULL OR status != 'Inactive')
           ORDER BY date_joined"""
        , (six_months_ago,)
    ).fetchall()

    confirmed = db.execute(
        """SELECT emp_id, full_name, department, position, confirmation_date, confirmed_new_salary
           FROM employees WHERE confirmation_date IS NOT NULL AND confirmation_date != ''
           ORDER BY confirmation_date DESC"""
    ).fetchall()

    # Recent self-service profile updates, most recent first - viewing this
    # page marks them all seen (clears the Alerts nav badge), like reading
    # a notification inbox.
    profile_updates = db.execute(
        """SELECT pul.id, pul.emp_id, pul.updated_at, e.full_name FROM profile_update_log pul
           JOIN employees e ON e.emp_id = pul.emp_id
           ORDER BY pul.updated_at DESC LIMIT 50"""
    ).fetchall()
    db.execute(
        "UPDATE profile_update_log SET hr_viewed_at=? WHERE hr_viewed_at IS NULL",
        (datetime.datetime.now().isoformat(timespec="seconds"),),
    )
    db.commit()

    return render_template("confirmation_due.html", due=due, missing=missing, today=today,
                            passport_alerts=passport_alerts, work_permit_alerts=work_permit_alerts,
                            confirmed=confirmed, profile_updates=profile_updates)


@app.route("/confirmation-letter/<emp_id>")
def confirmation_letter(emp_id):
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None:
        return "Employee not found", 404

    letter_date_str = request.args.get("letter_date")
    confirm_date_str = (
        request.args.get("confirm_date")
        or emp["confirmation_date"]
        or emp["probation_end_date"]
    )

    today = datetime.date.today()
    letter_date = datetime.date.fromisoformat(letter_date_str) if letter_date_str else today
    confirm_date = datetime.date.fromisoformat(confirm_date_str) if confirm_date_str else today

    hist_row = db.execute(
        """SELECT old_salary, new_salary FROM salary_history
           WHERE emp_id=? AND reason='Confirmation' AND effective_date=?
           ORDER BY recorded_at DESC LIMIT 1""",
        (emp_id, confirm_date.isoformat()),
    ).fetchone()

    new_salary_str = request.args.get("new_salary")
    if new_salary_str:
        new_salary = float(new_salary_str)
    elif hist_row:
        new_salary = hist_row["new_salary"]
    elif emp["confirmed_new_salary"]:
        new_salary = emp["confirmed_new_salary"]
    else:
        new_salary = None
    increment = None
    if new_salary is not None:
        # Basic Salary on the employee record already reflects the *new*
        # salary once Update Salary has been clicked, so the previous salary
        # for this letter has to come from the matching history row instead.
        old_salary = hist_row["old_salary"] if hist_row else emp["basic_salary"]
        increment = {
            "new_salary": new_salary,
            "old_salary": old_salary,
            "amount": round(new_salary - old_salary, 2),
            "pct": round((new_salary - old_salary) / old_salary * 100, 1)
                   if old_salary else None,
        }

    return render_template(
        "confirmation_letter.html", emp=emp,
        letter_date=letter_date, confirm_date=confirm_date, increment=increment,
    )


@app.route("/confirmation-appraisal/<emp_id>")
def confirmation_appraisal(emp_id):
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None:
        return "Employee not found", 404

    appraisal_date_str = request.args.get("appraisal_date")
    appraisal_date = (
        datetime.date.fromisoformat(appraisal_date_str) if appraisal_date_str
        else datetime.date.today()
    )

    # Leave & Discipline Records: summed from Attendance across every month
    # from Date Joined through the appraisal date - this app doesn't track
    # warning letters / domestic inquiries / late-ins, so those stay blank
    # for HR to fill in by hand.
    leave_totals = {"mc_days": 0, "ul_days": 0, "absent_days": 0}
    record_from = emp["date_joined"]
    if record_from:
        join_year, join_month = (int(p) for p in record_from.split("-")[:2])
        row = db.execute(
            """SELECT COALESCE(SUM(mc_days),0) AS mc_days, COALESCE(SUM(ul_days),0) AS ul_days,
                      COALESCE(SUM(absent_days),0) AS absent_days
               FROM attendance_monthly WHERE emp_id=?
                 AND (year > ? OR (year = ? AND month >= ?))
                 AND (year < ? OR (year = ? AND month <= ?))""",
            (emp_id, join_year, join_year, join_month,
             appraisal_date.year, appraisal_date.year, appraisal_date.month),
        ).fetchone()
        leave_totals = dict(row)

    return render_template(
        "confirmation_appraisal.html", emp=emp, appraisal_date=appraisal_date,
        record_from=record_from, leave_totals=leave_totals,
    )


@app.route("/resignation-letter/<emp_id>")
def resignation_letter(emp_id):
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None:
        return "Employee not found", 404

    today = datetime.date.today()

    letter_date_str = request.args.get("letter_date")
    letter_date = datetime.date.fromisoformat(letter_date_str) if letter_date_str else today

    resign_date_str = request.args.get("resignation_date") or emp["resignation_date"]
    resignation_date = datetime.date.fromisoformat(resign_date_str) if resign_date_str else None

    lwd_str = request.args.get("last_working_day") or emp["last_working_day"]
    last_working_day = datetime.date.fromisoformat(lwd_str) if lwd_str else None

    notice_days = None
    if resignation_date and last_working_day:
        notice_days = (last_working_day - resignation_date).days

    def _num(name):
        raw = request.args.get(name)
        return float(raw) if raw else None

    final_salary = _num("final_salary")
    notice_pay = _num("notice_pay")
    leave_days = _num("leave_days")
    deduction_amount = _num("deduction_amount") or 0.0
    deduction_desc = request.args.get("deduction_desc") or ""

    settlement = None
    if final_salary is not None or notice_pay is not None or leave_days is not None or deduction_amount:
        daily_rate = (emp["basic_salary"] or 0) / 26
        leave_amount = round((leave_days or 0) * daily_rate, 2)
        net_amount = round(
            (final_salary or 0) + (notice_pay or 0) + leave_amount - deduction_amount, 2
        )
        settlement = {
            "final_salary": final_salary,
            "notice_pay": notice_pay,
            "leave_days": leave_days,
            "daily_rate": round(daily_rate, 2),
            "leave_amount": leave_amount,
            "deduction_amount": deduction_amount,
            "deduction_desc": deduction_desc,
            "net_amount": net_amount,
        }

    return render_template(
        "resignation_letter.html", emp=emp, letter_date=letter_date,
        resignation_date=resignation_date, last_working_day=last_working_day,
        notice_days=notice_days, settlement=settlement,
    )


@app.route("/cp22a/<emp_id>")
def cp22a_form(emp_id):
    """Reference/data-collection sheet for LHDN's CP22A ('Notification by
    Employer of Cessation of Employment'). Since 1 Jan 2024 LHDN requires
    this to be filed online via the e-SPC service in MyTax
    (mytax.hasil.gov.my), not mailed as a paper form - so this page isn't a
    submittable form itself, just every figure HR needs gathered in one
    place (YTD remuneration/EPF/PCB pulled straight from payroll) to
    transcribe into e-SPC accurately, plus a checklist of what's still
    missing on file (e.g. employee's own tax no., forwarding address)."""
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None:
        return "Employee not found", 404

    cessation_date_str = request.args.get("cessation_date") or emp["last_working_day"] or emp["resignation_date"]
    cessation_date = datetime.date.fromisoformat(cessation_date_str) if cessation_date_str else None
    year = cessation_date.year if cessation_date else datetime.date.today().year

    monthly = db.execute(
        """SELECT month, gross_remun, epf_employee, pcb_deducted
           FROM pcb_monthly_record WHERE emp_id=? AND year=? ORDER BY month""",
        (emp_id, year),
    ).fetchall()
    totals = {
        "gross_remun": round(sum(m["gross_remun"] or 0 for m in monthly), 2),
        "epf_employee": round(sum(m["epf_employee"] or 0 for m in monthly), 2),
        "pcb_deducted": round(sum(m["pcb_deducted"] or 0 for m in monthly), 2),
    }

    return render_template(
        "cp22a_form.html", emp=emp, year=year, cessation_date=cessation_date,
        monthly=monthly, totals=totals, today=datetime.date.today(),
    )


@app.route("/hr/pcb-history-template")
def pcb_history_template():
    """Downloadable Excel template for bulk-loading pre-go-live PCB
    year-to-date history (Gross Remuneration, EPF Employee, PCB Deducted)
    per employee per month - needed whenever a company starts using this
    system mid-year, so PCB from the first live month onward is computed
    against accurate YTD figures under LHDN's Computerised Method, and so
    CP22A cessation notices show the full year's figures, not just the
    months actually run in this system."""
    db = get_db()
    employees = db.execute(
        "SELECT emp_id, full_name FROM employees ORDER BY emp_id"
    ).fetchall()
    year = request.args.get("year", type=int) or datetime.date.today().year
    upto_month = request.args.get("upto_month", type=int) or 12

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PCB History"
    ws.append(["YearMonth", "Emp ID", "Full Name (reference only)",
               "Gross Remuneration (RM)", "EPF Employee (RM)", "PCB Deducted (RM)"])
    for month in range(1, upto_month + 1):
        yearmonth = year * 100 + month
        for emp in employees:
            ws.append([yearmonth, emp["emp_id"], emp["full_name"], None, None, None])
    for col_idx, width in enumerate([12, 10, 32, 20, 16, 16], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=PCB_History_Template_{year}.xlsx"},
    )


@app.route("/hr/pcb-history-import", methods=["GET", "POST"])
def pcb_history_import():
    error = None
    imported = []
    skipped = []
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            error = "Please choose a file to upload."
        else:
            try:
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                db = get_db()
                valid_emp_ids = {r["emp_id"] for r in db.execute("SELECT emp_id FROM employees").fetchall()}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None or row[1] is None:
                        continue
                    yearmonth_raw, emp_id = row[0], str(row[1]).strip()
                    gross, epf, pcb = row[3], row[4], row[5]
                    if gross is None and epf is None and pcb is None:
                        continue  # left blank - not filled in, skip silently
                    if emp_id not in valid_emp_ids:
                        skipped.append(f"{emp_id} / {yearmonth_raw}: unknown Emp ID")
                        continue
                    try:
                        yearmonth = int(yearmonth_raw)
                        year, month = yearmonth // 100, yearmonth % 100
                        if not 1 <= month <= 12:
                            raise ValueError
                    except (TypeError, ValueError):
                        skipped.append(f"{emp_id} / {yearmonth_raw}: invalid YearMonth")
                        continue
                    db.execute(
                        """INSERT INTO pcb_monthly_record (emp_id, year, month, gross_remun, epf_employee, pcb_deducted)
                           VALUES (?,?,?,?,?,?)
                           ON CONFLICT(emp_id, year, month) DO UPDATE SET
                             gross_remun=excluded.gross_remun, epf_employee=excluded.epf_employee,
                             pcb_deducted=excluded.pcb_deducted""",
                        (emp_id, year, month, float(gross or 0), float(epf or 0), float(pcb or 0)),
                    )
                    imported.append(f"{emp_id} {year}-{month:02d}: Gross RM{float(gross or 0):.2f}, "
                                     f"EPF RM{float(epf or 0):.2f}, PCB RM{float(pcb or 0):.2f}")
                db.commit()
            except Exception as exc:
                error = f"Could not read file: {exc}"
    return render_template("pcb_history_import.html", error=error, imported=imported, skipped=skipped)


@app.route("/leave-form")
def leave_application_form():
    db = get_db()
    emp = None
    emp_id = request.args.get("emp_id")
    if emp_id:
        emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()

    employees = db.execute(
        "SELECT emp_id, full_name FROM employees WHERE status != 'Inactive' ORDER BY emp_id"
    ).fetchall()

    balances = None
    if emp:
        year = datetime.date.today().year
        totals = db.execute(
            """SELECT COALESCE(SUM(al_days),0) AS al, COALESCE(SUM(mc_days),0) AS mc,
                      COALESCE(SUM(hl_days),0) AS hl
               FROM attendance_monthly WHERE emp_id=? AND year=?""",
            (emp_id, year),
        ).fetchone()
        prorated_al, _al_note = _prorated_al_note(emp, year)
        al_entitlement = prorated_al if prorated_al is not None else (emp["annual_leave_entitlement"] or 0)
        balances = {
            "al": al_entitlement - totals["al"],
            "mc": (emp["mc_entitlement"] or 0) - totals["mc"],
            "hl": (emp["hospitalisation_leave_entitlement"] or 0) - totals["hl"],
        }

    return render_template(
        "leave_application_form.html", emp=emp, employees=employees, balances=balances,
        today=datetime.date.today(),
    )


# ---------------- Payroll ----------------

@app.route("/payroll/<int:year>/<int:month>")
def run_payroll(year, month):
    db = get_db()
    emps = employed_this_month(db, year, month)
    results = [payroll_calc.get_payroll_result(db, r["emp_id"], year, month) for r in emps]
    finalized = {
        r["emp_id"] for r in db.execute(
            "SELECT emp_id FROM payroll_runs WHERE year=? AND month=?", (year, month)
        ).fetchall()
    }
    bank_info = {
        r["emp_id"]: {"bank_name": r["bank_name"], "bank_account_no": r["bank_account_no"]}
        for r in db.execute("SELECT emp_id, bank_name, bank_account_no FROM employees").fetchall()
    }
    totals = {
        k: round(sum(r[k] for r in results), 2)
        for k in ["basic_salary", "unpaid_deduction", "gross_pay", "epf_employee", "epf_employer",
                   "socso_employee", "socso_employer", "eis_employee", "eis_employer", "pcb",
                   "skbbk_employee", "hrd_levy_employer", "other_deduction",
                   "total_deductions", "net_pay"]
    }
    return render_template("payroll.html", results=results, year=year, month=month,
                            finalized=finalized, totals=totals, bank_info=bank_info,
                            zero_pay_notes=_zero_pay_notes(results))


PAYROLL_EXPORT_COLUMNS = [
    ("Emp ID", "emp_id"), ("Name", "full_name"), ("Basic", "basic"),
    ("Fixed Allow.", "fixed_allowance"), ("Var. Allow.", "variable_allowance"),
    ("OT 1.5h", "ot_hours_1_5"), ("OT 2.0h", "ot_hours_2_0"), ("OT 3.0h", "ot_hours_3_0"),
    ("OT Rate", "ot_hourly_rate"), ("Total OT", "ot_pay"), ("Transport", "transport_allowance"),
    ("Meal", "meal_allowance"), ("CEWI", "cewi_allowance"), ("Gross", "gross"),
    ("EPF (Emp)", "epf_employee"), ("SOCSO (Emp)", "socso_employee"),
    ("EIS (Emp)", "eis_employee"), ("SKBBK", "skbbk_employee"), ("PCB", "pcb"),
    ("EPF (Er)", "epf_employer"), ("SOCSO (Er)", "socso_employer"),
    ("EIS (Er)", "eis_employer"), ("HRD Levy", "hrd_levy_employer"),
    ("UL Deduction", "unpaid_deduction"), ("Other Ded.", "other_deduction"),
    ("Total Ded.", "total_deduction"), ("NET PAY", "net_pay"), ("Status", "status"),
    ("Bank", "bank_name"), ("Acct No.", "bank_account_no"),
]


@app.route("/payroll/<int:year>/<int:month>/export")
def payroll_export(year, month):
    """Same figures/columns shown on the Run Payroll page, as a downloadable
    .xlsx - built from the same get_payroll_result() calls so it can never
    drift from what's on screen (frozen for finalized months, live preview
    otherwise, exactly like the page itself)."""
    db = get_db()
    emps = employed_this_month(db, year, month)
    results = [payroll_calc.get_payroll_result(db, r["emp_id"], year, month) for r in emps]
    finalized = {
        r["emp_id"] for r in db.execute(
            "SELECT emp_id FROM payroll_runs WHERE year=? AND month=?", (year, month)
        ).fetchall()
    }
    bank_info = {
        r["emp_id"]: {"bank_name": r["bank_name"], "bank_account_no": r["bank_account_no"]}
        for r in db.execute("SELECT emp_id, bank_name, bank_account_no FROM employees").fetchall()
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{MONTH_NAMES[month]} {year}"[:31]

    ws.cell(row=1, column=1, value=f"Payroll: {year:04d}{month:02d} End Month").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PAYROLL_EXPORT_COLUMNS))

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for col_idx, (label, _) in enumerate(PAYROLL_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill

    money_cols = {"Basic", "Fixed Allow.", "Var. Allow.", "Total OT", "Transport", "Meal", "CEWI",
                  "Gross", "EPF (Emp)", "EPF (Er)", "SOCSO (Emp)", "SOCSO (Er)", "EIS (Emp)",
                  "EIS (Er)", "PCB", "SKBBK", "HRD Levy", "UL Deduction", "Other Ded.",
                  "Total Ded.", "NET PAY"}

    row_idx = 3
    for r in results:
        row = {
            "emp_id": r["emp_id"], "full_name": r["full_name"],
            "basic": round(r["basic_salary"] + (r["unpaid_deduction"] or 0), 2),
            "fixed_allowance": r["fixed_allowance"], "variable_allowance": r["variable_allowance"],
            "ot_hours_1_5": r["ot_hours_1_5"], "ot_hours_2_0": r["ot_hours_2_0"],
            "ot_hours_3_0": r["ot_hours_3_0"], "ot_hourly_rate": round(r["ot_hourly_rate"], 4),
            "ot_pay": r["ot_pay"], "transport_allowance": r["transport_allowance"],
            "meal_allowance": r["meal_allowance"], "cewi_allowance": r["cewi_allowance"],
            "gross": round(r["gross_pay"] + (r["unpaid_deduction"] or 0), 2),
            "epf_employee": r["epf_employee"], "epf_employer": r["epf_employer"],
            "socso_employee": r["socso_employee"], "socso_employer": r["socso_employer"],
            "eis_employee": r["eis_employee"], "eis_employer": r["eis_employer"], "pcb": r["pcb"],
            "skbbk_employee": r["skbbk_employee"], "hrd_levy_employer": r["hrd_levy_employer"],
            "unpaid_deduction": r["unpaid_deduction"] or 0, "other_deduction": r["other_deduction"],
            "total_deduction": round(r["total_deductions"] + (r["unpaid_deduction"] or 0), 2),
            "net_pay": r["net_pay"],
            "status": "Finalized" if r["emp_id"] in finalized else "Draft",
            "bank_name": bank_info.get(r["emp_id"], {}).get("bank_name") or "",
            "bank_account_no": bank_info.get(r["emp_id"], {}).get("bank_account_no") or "",
        }
        for col_idx, (label, key) in enumerate(PAYROLL_EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            if label in money_cols:
                cell.number_format = "#,##0.00"
        row_idx += 1

    totals = {
        k: round(sum(r[k] for r in results), 2)
        for k in ["gross_pay", "epf_employee", "epf_employer", "socso_employee", "socso_employer",
                   "eis_employee", "eis_employer", "pcb", "skbbk_employee", "hrd_levy_employer",
                   "unpaid_deduction", "other_deduction", "total_deductions", "net_pay"]
    }
    total_row = {
        "emp_id": "TOTAL", "full_name": "", "basic": "", "fixed_allowance": "",
        "variable_allowance": "", "ot_hours_1_5": "", "ot_hours_2_0": "", "ot_hours_3_0": "",
        "ot_hourly_rate": "", "ot_pay": "", "transport_allowance": "", "meal_allowance": "",
        "cewi_allowance": "", "gross": round(totals["gross_pay"] + totals["unpaid_deduction"], 2),
        "epf_employee": totals["epf_employee"], "epf_employer": totals["epf_employer"],
        "socso_employee": totals["socso_employee"], "socso_employer": totals["socso_employer"],
        "eis_employee": totals["eis_employee"], "eis_employer": totals["eis_employer"],
        "pcb": totals["pcb"], "skbbk_employee": totals["skbbk_employee"],
        "hrd_levy_employer": totals["hrd_levy_employer"], "unpaid_deduction": totals["unpaid_deduction"],
        "other_deduction": totals["other_deduction"],
        "total_deduction": round(totals["total_deductions"] + totals["unpaid_deduction"], 2),
        "net_pay": totals["net_pay"], "status": "", "bank_name": "", "bank_account_no": "",
    }
    for col_idx, (label, key) in enumerate(PAYROLL_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=total_row[key])
        cell.font = Font(bold=True)
        if label in money_cols and total_row[key] != "":
            cell.number_format = "#,##0.00"

    last_row = _add_headcount_block(ws, row_idx + 2, totals, len(results))
    _add_zero_pay_notes(ws, last_row + 2, _zero_pay_notes(results))

    ws.freeze_panes = "C3"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    for col_idx in range(3, len(PAYROLL_EXPORT_COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
    _set_a4_one_page(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Payroll_{MONTH_NAMES[month]}_{year}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _payroll_summary_data(db, year, month):
    emps = employed_this_month(db, year, month)
    results = [payroll_calc.get_payroll_result(db, r["emp_id"], year, month) for r in emps]
    bank_info = {
        r["emp_id"]: {"bank_name": r["bank_name"], "bank_account_no": r["bank_account_no"]}
        for r in db.execute("SELECT emp_id, bank_name, bank_account_no FROM employees").fetchall()
    }
    for r in results:
        r["basic"] = round(r["basic_salary"] + (r["unpaid_deduction"] or 0), 2)
        r["total_allowance"] = round(
            (r["fixed_allowance"] or 0) + (r["variable_allowance"] or 0)
            + (r["transport_allowance"] or 0) + (r["meal_allowance"] or 0) + (r["cewi_allowance"] or 0), 2)
        r["bank_name"] = bank_info.get(r["emp_id"], {}).get("bank_name") or ""
        r["bank_account_no"] = bank_info.get(r["emp_id"], {}).get("bank_account_no") or ""
    totals = {
        k: round(sum(r[k] for r in results), 2)
        for k in ["basic", "total_allowance", "ot_pay", "gross_pay", "epf_employee", "epf_employer",
                   "socso_employee", "socso_employer", "eis_employee", "eis_employer", "pcb",
                   "skbbk_employee", "hrd_levy_employer", "total_deductions", "net_pay"]
    }
    return results, totals


def _zero_pay_notes(results):
    """Flags anyone whose Net Pay is RM0 (or less) because Unpaid Leave/Absent
    covered the whole month, so a report reader doesn't mistake it for a
    data error. Data-driven off whatever's actually in payroll this run,
    not a fixed list of names/months."""
    notes = []
    for r in results:
        if (r["net_pay"] or 0) <= 0 and (r["unpaid_days"] or 0) > 0:
            notes.append(
                f"{r['emp_id']} ({r['full_name']}) shows RM0.00 pay due to "
                f"{r['unpaid_days']:g} day(s) Unpaid Leave/Absence this month."
            )
    return notes


@app.route("/payroll/<int:year>/<int:month>/summary")
def payroll_summary(year, month):
    results, totals = _payroll_summary_data(get_db(), year, month)
    totals_only = request.args.get("totals_only") == "1"
    return render_template("payroll_summary.html", results=results, totals_only=totals_only,
                            year=year, month=month, totals=totals,
                            zero_pay_notes=_zero_pay_notes(results))


@app.route("/payroll/<int:year>/<int:month>/summary/view")
def payroll_summary_view(year, month):
    results, totals = _payroll_summary_data(get_db(), year, month)
    totals_only = request.args.get("totals_only") == "1"
    return render_template("payroll_summary_view.html", results=results, totals_only=totals_only,
                            year=year, month=month, totals=totals,
                            zero_pay_notes=_zero_pay_notes(results))


PAYROLL_SUMMARY_EXPORT_COLUMNS = [
    ("Emp ID", "emp_id"), ("Name", "full_name"), ("Basic", "basic"),
    ("Total Allowance", "total_allowance"),
    ("Total OT", "ot_pay"), ("Gross Pay", "gross_pay"),
    ("EPF (Emp)", "epf_employee"), ("SOCSO (Emp)", "socso_employee"),
    ("EIS (Emp)", "eis_employee"), ("SKBBK", "skbbk_employee"), ("PCB", "pcb"),
    ("EPF (Er)", "epf_employer"), ("SOCSO (Er)", "socso_employer"),
    ("EIS (Er)", "eis_employer"), ("HRD Levy", "hrd_levy_employer"),
    ("Total Ded.", "total_deductions"),
    ("NET PAY", "net_pay"), ("Bank", "bank_name"), ("Acct No.", "bank_account_no"),
]


@app.route("/payroll/<int:year>/<int:month>/summary/export")
def payroll_summary_export(year, month):
    db = get_db()
    results, totals = _payroll_summary_data(db, year, month)
    totals_only = request.args.get("totals_only") == "1"
    non_aggregate_keys = ("emp_id", "full_name", "bank_name", "bank_account_no")
    columns = (
        [c for c in PAYROLL_SUMMARY_EXPORT_COLUMNS if c[1] not in non_aggregate_keys]
        if totals_only else PAYROLL_SUMMARY_EXPORT_COLUMNS
    )
    money_cols = {label for label, key in PAYROLL_SUMMARY_EXPORT_COLUMNS if key not in non_aggregate_keys}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{MONTH_NAMES[month]} {year} Summary"[:31]

    ws.cell(row=1, column=1, value=f"Payroll Summary: {year:04d}{month:02d} End Month").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for col_idx, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill

    row_idx = 3
    if not totals_only:
        for r in results:
            for col_idx, (label, key) in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=r[key])
                if label in money_cols:
                    cell.number_format = "#,##0.00"
            row_idx += 1

    label_col_span = 1 if totals_only else 2
    total_label = f"TOTAL ({len(results)} employees)"
    ws.cell(row=row_idx, column=1, value=total_label).font = Font(bold=True)
    for col_idx, (label, key) in enumerate(columns, start=1):
        if col_idx <= label_col_span or key in ("bank_name", "bank_account_no"):
            continue
        cell = ws.cell(row=row_idx, column=col_idx, value=totals[key])
        cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"

    last_row = _add_headcount_block(ws, row_idx + 2, totals, len(results))
    _add_zero_pay_notes(ws, last_row + 2, _zero_pay_notes(results))

    if not totals_only:
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 26
    for col_idx in range(label_col_span + 1, len(columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14
    _set_a4_one_page(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffix = "_TotalsOnly" if totals_only else ""
    filename = f"Payroll_Summary_{MONTH_NAMES[month]}_{year}{suffix}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/payroll/<int:year>/<int:month>/finalize", methods=["POST"])
def finalize_payroll(year, month):
    db = get_db()
    emps = employed_this_month(db, year, month)
    now = datetime.datetime.now().isoformat(timespec="seconds")
    for r in emps:
        result = payroll_calc.calculate_payroll(db, r["emp_id"], year, month)
        db.execute(
            """INSERT INTO payroll_runs (
                emp_id, year, month, basic_salary, fixed_allowance, variable_allowance,
                working_days_in_month, days_worked, paid_leave_days, unpaid_days, unpaid_deduction,
                ot_hours_1_5, ot_hours_2_0, ot_hours_3_0, ot_pay_1_5, ot_pay_2_0, ot_pay_3_0,
                ot_hourly_rate, ot_pay, days_employed, prorate_factor, transport_allowance,
                meal_allowance, cewi_allowance, gross_pay, epf_employee, additional_epf_employee,
                epf_employer, socso_employee, socso_employer, eis_employee, eis_employer, pcb,
                skbbk_employee, hrd_levy_employer, other_deduction, other_deduction_desc,
                total_deductions, net_pay, finalized_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(emp_id, year, month) DO UPDATE SET
                basic_salary=excluded.basic_salary, fixed_allowance=excluded.fixed_allowance,
                variable_allowance=excluded.variable_allowance,
                working_days_in_month=excluded.working_days_in_month,
                days_worked=excluded.days_worked, paid_leave_days=excluded.paid_leave_days,
                unpaid_days=excluded.unpaid_days, unpaid_deduction=excluded.unpaid_deduction,
                ot_hours_1_5=excluded.ot_hours_1_5, ot_hours_2_0=excluded.ot_hours_2_0,
                ot_hours_3_0=excluded.ot_hours_3_0, ot_pay_1_5=excluded.ot_pay_1_5,
                ot_pay_2_0=excluded.ot_pay_2_0, ot_pay_3_0=excluded.ot_pay_3_0,
                ot_hourly_rate=excluded.ot_hourly_rate, ot_pay=excluded.ot_pay,
                days_employed=excluded.days_employed, prorate_factor=excluded.prorate_factor,
                transport_allowance=excluded.transport_allowance,
                meal_allowance=excluded.meal_allowance, cewi_allowance=excluded.cewi_allowance,
                gross_pay=excluded.gross_pay, net_pay=excluded.net_pay,
                pcb=excluded.pcb, epf_employee=excluded.epf_employee,
                additional_epf_employee=excluded.additional_epf_employee,
                epf_employer=excluded.epf_employer, socso_employee=excluded.socso_employee,
                socso_employer=excluded.socso_employer, eis_employee=excluded.eis_employee,
                eis_employer=excluded.eis_employer, skbbk_employee=excluded.skbbk_employee,
                hrd_levy_employer=excluded.hrd_levy_employer,
                other_deduction=excluded.other_deduction, other_deduction_desc=excluded.other_deduction_desc,
                total_deductions=excluded.total_deductions, finalized_at=excluded.finalized_at""",
            (
                result["emp_id"], year, month, result["basic_salary"], result["fixed_allowance"],
                result["variable_allowance"], result["working_days_in_month"],
                result["working_days_in_month"] - result["unpaid_days"] - result["paid_leave_days"],
                result["paid_leave_days"], result["unpaid_days"], result["unpaid_deduction"],
                result["ot_hours_1_5"], result["ot_hours_2_0"], result["ot_hours_3_0"],
                result["ot_pay_1_5"], result["ot_pay_2_0"], result["ot_pay_3_0"],
                result["ot_hourly_rate"], result["ot_pay"], result["days_employed"], result["prorate_factor"],
                result["transport_allowance"], result["meal_allowance"], result["cewi_allowance"],
                result["gross_pay"], result["epf_employee"], result["additional_epf_employee"],
                result["epf_employer"],
                result["socso_employee"], result["socso_employer"], result["eis_employee"],
                result["eis_employer"], result["pcb"], result["skbbk_employee"],
                result["hrd_levy_employer"], result["other_deduction"], result["other_deduction_desc"],
                result["total_deductions"], result["net_pay"], now,
            ),
        )
        db.execute(
            """INSERT INTO pcb_monthly_record (emp_id, year, month, gross_remun, epf_employee, pcb_deducted)
               VALUES (?,?,?,?,?,?)
               ON CONFLICT(emp_id, year, month) DO UPDATE SET
                 gross_remun=excluded.gross_remun, epf_employee=excluded.epf_employee,
                 pcb_deducted=excluded.pcb_deducted""",
            (result["emp_id"], year, month, result["gross_pay"], result["epf_employee"], result["pcb"]),
        )
    db.commit()
    return redirect(url_for("run_payroll", year=year, month=month))


# PERKESO's official fixed-width text file layout for combined SOCSO + EIS +
# SKBBK contribution submission via the ASSIST Portal (v2.0, 13 Feb 2026 -
# https://www.perkeso.gov.my/images/panduan/280722_-_TEXTFILE_STRUCTURE_FOR_COMBINE_CONTRIBUTION_V1.pdf).
# One 278-character line per employee, no header/footer, plain ASCII.
# Numeric fields are amounts in CENTS (no decimal point), right-justified,
# zero-padded; alphanumeric fields are left-justified, space-padded.
SOCSO_EIS_EMPLOYER_MYCOID = "202301036641"  # Company's registered no. (see letterhead)
SOCSO_EIS_EMPLOYER_CODE = "E1102131652V"     # PERKESO Employer Code, from the ASSIST Portal


def _pad_left(value, width):
    """Alphanumeric field: left-justified, space-padded, truncated if too long."""
    return str(value or "")[:width].ljust(width)


def _pad_num(cents, width):
    """Numeric field: right-justified, zero-padded amount in cents."""
    return str(int(round(cents)))[:width].rjust(width, "0")


@app.route("/socso-eis-textfile/<int:year>/<int:month>")
def socso_eis_textfile(year, month):
    db = get_db()
    rows = db.execute(
        """SELECT pr.*, e.ic_passport_no, e.full_name
           FROM payroll_runs pr JOIN employees e ON e.emp_id = pr.emp_id
           WHERE pr.year=? AND pr.month=? AND pr.gross_pay > 0 ORDER BY pr.emp_id""",
        (year, month),
    ).fetchall()

    month_str = f"{month:02d}{year:04d}"  # MMYYYY
    lines = []
    for r in rows:
        line = (
            _pad_left(SOCSO_EIS_EMPLOYER_CODE, 12)   # 1. Employer Code
            + _pad_left(SOCSO_EIS_EMPLOYER_MYCOID, 20)  # 2. MyCoID / SSM No.
            + _pad_left(r["ic_passport_no"], 12)     # 3. ID No. / SOCSO Foreign Worker No.
            + _pad_left(r["full_name"], 150)         # 4. Employee Name
            + month_str                              # 5. Month Contribution (MMYYYY)
            + _pad_num((r["gross_pay"] or 0) * 100, 14)       # 6. Employee Salary (cents)
            + _pad_num((r["socso_employer"] or 0) * 100, 6)   # 7. SOCSO - Employer (cents)
            + _pad_num((r["socso_employee"] or 0) * 100, 6)   # 8. SOCSO - Employee (cents)
            + _pad_num((r["eis_employer"] or 0) * 100, 6)     # 9. EIS - Employer (cents)
            + _pad_num((r["eis_employee"] or 0) * 100, 6)     # 10. EIS - Employee (cents)
            + _pad_num((r["skbbk_employee"] or 0) * 100, 6)   # 11. SKBBK - Employee (cents)
            + _pad_left("", 14)                      # 12. Filler 1
            + _pad_left("", 20)                       # 13. Filler 2
        )
        lines.append(line)

    content = "\r\n".join(lines) + "\r\n" if lines else ""
    filename = f"SOCSO_EIS_{month_str}.txt"
    return Response(
        content, mimetype="text/plain",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/payslip/<emp_id>/<int:year>/<int:month>")
def payslip(emp_id, year, month):
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    result = payroll_calc.get_payroll_result(db, emp_id, year, month)
    return render_template("payslip.html", emp=emp, r=result, year=year, month=month,
                            payment_date=payment_date_for(year, month))


@app.route("/payroll/<int:year>/<int:month>/<emp_id>/detail")
def payroll_detail(emp_id, year, month):
    """Formula-level breakdown of how every figure was calculated (which
    wage bracket was matched, the unpaid-leave/PCB formulas with actual
    numbers plugged in, etc.) - not just the final RM amounts already shown
    on the Payslip. Always computed live from calculate_payroll() (the same
    single engine the Payslip/Finalize routes use, so the explanation can
    never drift out of sync with the real logic); if this month was already
    finalized, its frozen net pay is compared against today's live figure
    so a mismatch (e.g. attendance edited after finalizing) is surfaced
    honestly instead of silently shown as if it were the saved amount."""
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None:
        return "Employee not found", 404

    live = payroll_calc.calculate_payroll(db, emp_id, year, month)
    frozen_row = db.execute(
        "SELECT * FROM payroll_runs WHERE emp_id=? AND year=? AND month=?",
        (emp_id, year, month),
    ).fetchone()
    is_finalized = frozen_row is not None
    mismatch = is_finalized and abs((frozen_row["net_pay"] or 0) - live["net_pay"]) > 0.01

    return render_template(
        "payroll_detail.html", emp=emp, r=live, year=year, month=month,
        is_finalized=is_finalized, mismatch=mismatch,
        frozen_net_pay=(frozen_row["net_pay"] if frozen_row else None),
    )


@app.route("/payroll/<int:year>/<int:month>/<emp_id>/pcb-override", methods=["POST"])
def set_pcb_override(year, month, emp_id):
    """Manually corrects PCB for one finalized employee/month, e.g. to match
    the employee's own official LHDN e-PCB slip when this employer's
    Computerised Method doesn't exactly reproduce LHDN's figure (edge cases
    like irregular voluntary EPF elections). Only valid for an already
    finalized month - the override is stored on its payroll_runs row and
    picked up by calculate_payroll() from then on, including through future
    re-Finalizes."""
    db = get_db()
    existing = db.execute(
        "SELECT 1 FROM payroll_runs WHERE emp_id=? AND year=? AND month=?", (emp_id, year, month)
    ).fetchone()
    if existing is None:
        return "This month hasn't been finalized yet - finalize it first.", 400

    raw = request.form.get("pcb_override", "").strip()
    reason = request.form.get("pcb_override_reason", "").strip() or None
    override_value = float(raw) if raw else None

    db.execute(
        "UPDATE payroll_runs SET pcb_override=?, pcb_override_reason=? WHERE emp_id=? AND year=? AND month=?",
        (override_value, reason, emp_id, year, month),
    )
    # Recompute now so the corrected figure shows everywhere immediately,
    # without waiting for the next Finalize.
    result = payroll_calc.calculate_payroll(db, emp_id, year, month)
    db.execute(
        "UPDATE payroll_runs SET pcb=?, total_deductions=?, net_pay=? WHERE emp_id=? AND year=? AND month=?",
        (result["pcb"], result["total_deductions"], result["net_pay"], emp_id, year, month),
    )
    db.execute(
        """INSERT INTO pcb_monthly_record (emp_id, year, month, gross_remun, epf_employee, pcb_deducted)
           VALUES (?,?,?,?,?,?)
           ON CONFLICT(emp_id, year, month) DO UPDATE SET pcb_deducted=excluded.pcb_deducted""",
        (emp_id, year, month, result["gross_pay"], result["epf_employee"], result["pcb"]),
    )
    db.commit()
    return redirect(url_for("payroll_detail", year=year, month=month, emp_id=emp_id))


# ---------------- Settings ----------------

@app.route("/settings", methods=["GET", "POST"])
def payroll_settings_page():
    db = get_db()
    error = None
    if request.method == "POST":
        raw = request.form.get("payslip_release_day", "").strip()
        try:
            day = int(raw)
            if not 1 <= day <= 31:
                raise ValueError
        except ValueError:
            error = "Enter a whole number from 1 to 31."
        else:
            db.execute(
                """INSERT INTO payroll_settings (key, value) VALUES ('payslip_release_day', ?)
                   ON CONFLICT(key) DO UPDATE SET value=excluded.value""",
                (str(day),),
            )
            db.commit()

    row = db.execute("SELECT value FROM payroll_settings WHERE key='payslip_release_day'").fetchone()
    release_day = int(row["value"]) if row else PAYMENT_DAY
    return render_template("settings.html", release_day=release_day, error=error,
                            payment_day=PAYMENT_DAY)


# ---------------- Public Holidays ----------------

@app.route("/holidays", methods=["GET", "POST"])
def holidays():
    db = get_db()
    if request.method == "POST":
        date = request.form.get("date")
        name = request.form.get("name")
        remarks = request.form.get("remarks") or None
        if date and name:
            day = datetime.date.fromisoformat(date).strftime("%A")
            db.execute(
                "INSERT INTO public_holidays (date, day, name, remarks) VALUES (?,?,?,?)",
                (date, day, name, remarks),
            )
            db.commit()
        return redirect(url_for("holidays"))
    rows = db.execute("SELECT * FROM public_holidays ORDER BY date").fetchall()
    return render_template("holidays.html", holidays=rows)


@app.route("/holidays/<int:holiday_id>/delete", methods=["POST"])
def delete_holiday(holiday_id):
    db = get_db()
    db.execute("DELETE FROM public_holidays WHERE id=?", (holiday_id,))
    db.commit()
    return redirect(url_for("holidays"))


# ---------------- Export API (for the Staff Leave Portal) ----------------
# Read-only JSON feed the Staff Leave Portal's HR dashboard pulls from to
# publish a month's payroll figures as staff payslips. CORS is wide open
# since this only ever serves localhost-to-localhost during development.

@app.route("/api/payroll-export/<int:year>/<int:month>")
def api_payroll_export(year, month):
    from flask import jsonify

    db = get_db()
    emps = db.execute(
        """SELECT emp_id, full_name, department, position, status, ic_passport_no,
                  date_of_birth, marital_status, date_joined, last_working_day,
                  probation_end_date, confirmation_date
           FROM employees ORDER BY emp_id"""
    ).fetchall()
    finalized_ids = {
        r["emp_id"] for r in db.execute(
            "SELECT emp_id FROM payroll_runs WHERE year=? AND month=?", (year, month)
        ).fetchall()
    }

    employees_out = []
    for e in emps:
        p = payroll_calc.get_payroll_result(db, e["emp_id"], year, month)
        employees_out.append({
            "empId": e["emp_id"],
            "name": e["full_name"],
            "department": e["department"],
            "position": e["position"],
            "status": e["status"],
            "icPassport": e["ic_passport_no"],
            "dob": e["date_of_birth"],
            "maritalStatus": e["marital_status"],
            "dateJoined": e["date_joined"],
            "lastWorkingDay": e["last_working_day"],
            "probationEndDate": e["probation_end_date"],
            "confirmationDate": e["confirmation_date"],
            "finalized": e["emp_id"] in finalized_ids,
            "payroll": {
                "basic": p["basic_salary"],
                "fixedAllow": p["fixed_allowance"],
                "varAllow": p["variable_allowance"],
                "transport": p["transport_allowance"],
                "meal": p["meal_allowance"],
                "cewi": p["cewi_allowance"],
                "ot15h": p["ot_hours_1_5"],
                "ot20h": p["ot_hours_2_0"],
                "ot30h": p["ot_hours_3_0"],
                "otRate": p["ot_hourly_rate"],
                "totalOt": p["ot_pay"],
                "gross": p["gross_pay"],
                "epfEmp": p["epf_employee"],
                "epfEr": p["epf_employer"],
                "socsoEmp": p["socso_employee"],
                "socsoEr": p["socso_employer"],
                "eisEmp": p["eis_employee"],
                "eisEr": p["eis_employer"],
                "pcb": p["pcb"],
                "skbbk": p["skbbk_employee"],
                "hrdLevy": p["hrd_levy_employer"],
                "netPay": p["net_pay"],
            },
        })

    resp = jsonify({
        "year": year,
        "month": month,
        "monthName": MONTH_NAMES[month],
        "employees": employees_out,
    })
    resp.headers["Access-Control-Allow-Origin"] = "*"
    return resp


# ---------------- Staff Portal: Pages ----------------

@app.route("/portal")
@portal_login_required
def portal_dashboard():
    db = get_db()
    emp = current_portal_employee(db)
    today = datetime.date.today()

    recent_runs = db.execute(
        """SELECT year, month, net_pay FROM payroll_runs WHERE emp_id=?
           ORDER BY year DESC, month DESC LIMIT 12""",
        (emp["emp_id"],),
    ).fetchall()
    latest_run = next(
        (r for r in recent_runs if today >= payslip_release_date(db, r["year"], r["month"])), None
    )

    al_used = db.execute(
        """SELECT COALESCE(SUM(al_days), 0) AS used FROM attendance_monthly
           WHERE emp_id=? AND year=?""",
        (emp["emp_id"], today.year),
    ).fetchone()["used"]
    prorated_al, _al_note = _prorated_al_note(emp, today.year)
    al_entitlement = prorated_al if prorated_al is not None else (emp["annual_leave_entitlement"] or 0)
    al_balance = al_entitlement - al_used

    pending_leave = db.execute(
        """SELECT COUNT(*) AS n FROM leave_requests WHERE emp_id=? AND status='Pending'""",
        (emp["emp_id"],),
    ).fetchone()["n"]
    pending_trip = db.execute(
        """SELECT COUNT(*) AS n FROM business_trips WHERE emp_id=? AND status='Pending'""",
        (emp["emp_id"],),
    ).fetchone()["n"]
    pending_own_ot = 0
    if emp["ot_approval_required"] == "Y":
        pending_own_ot = db.execute(
            """SELECT COUNT(*) AS n FROM ot_claims WHERE emp_id=? AND status='Pending'""",
            (emp["emp_id"],),
        ).fetchone()["n"]

    # If this employee is also an HR/approver user (e.g. Mr Kee = K003 and
    # hr_users 'kee'), surface their own supervisor reminders here too, so
    # they don't need a separate HR login just to see them.
    supervisor = None
    if emp["hr_username"]:
        hr_user = db.execute(
            "SELECT username, can_approve_leave, can_approve_appraisal, can_approve_ot FROM hr_users WHERE username=?",
            (emp["hr_username"],),
        ).fetchone()
        if hr_user:
            team_pending_leave = 0
            team_pending_trip = 0
            if hr_user["can_approve_leave"] == "Y":
                team_pending_leave = db.execute(
                    """SELECT COUNT(*) AS c FROM leave_requests lr
                       JOIN employees e ON e.emp_id = lr.emp_id
                       WHERE lr.status='Pending' AND e.leave_approver_username=?""",
                    (hr_user["username"],),
                ).fetchone()["c"]
                team_pending_trip = db.execute(
                    """SELECT COUNT(*) AS c FROM business_trips bt
                       JOIN employees e ON e.emp_id = bt.emp_id
                       WHERE bt.status='Pending' AND e.leave_approver_username=?""",
                    (hr_user["username"],),
                ).fetchone()["c"]
            team_pending_appraisal = 0
            if hr_user["can_approve_appraisal"] == "Y":
                team_pending_appraisal = db.execute(
                    """SELECT COUNT(*) AS c FROM employees e
                       WHERE e.appraisal_supervisor_username=? AND e.status != 'Inactive'
                         AND NOT EXISTS (SELECT 1 FROM appraisals a WHERE a.emp_id = e.emp_id)""",
                    (hr_user["username"],),
                ).fetchone()["c"]
            # can_approve_ot is a single company-wide approver (not a per-
            # employee assignment like Leave Approver), so this counts every
            # pending OT claim, not just this supervisor's own team.
            team_pending_ot = 0
            if hr_user["can_approve_ot"] == "Y":
                team_pending_ot = db.execute(
                    "SELECT COUNT(*) AS c FROM ot_claims WHERE status='Pending'"
                ).fetchone()["c"]
            if hr_user["can_approve_leave"] == "Y" or hr_user["can_approve_appraisal"] == "Y" or hr_user["can_approve_ot"] == "Y":
                supervisor = {"pending_leave": team_pending_leave, "pending_trip": team_pending_trip,
                              "pending_appraisal": team_pending_appraisal, "pending_ot": team_pending_ot}

    return render_template(
        "portal_dashboard.html", emp=emp, latest_run=latest_run,
        al_balance=al_balance, pending_leave=pending_leave, pending_trip=pending_trip,
        pending_own_ot=pending_own_ot, supervisor=supervisor,
    )


@app.route("/portal/payslips")
@portal_login_required
def portal_payslips():
    db = get_db()
    emp = current_portal_employee(db)
    today = datetime.date.today()
    all_rows = db.execute(
        """SELECT year, month, net_pay, finalized_at FROM payroll_runs
           WHERE emp_id=? ORDER BY year DESC, month DESC""",
        (emp["emp_id"],),
    ).fetchall()
    # Only months whose release date has arrived are shown at all - a row
    # (and its Net Pay figure) simply doesn't exist yet from the employee's
    # point of view before then, not just the itemized breakdown.
    rows = [r for r in all_rows if today >= payslip_release_date(db, r["year"], r["month"])]
    # A payroll_runs row can exist for a month before someone actually
    # joined (e.g. finalized while their Date Joined was still wrong/unset,
    # later corrected to net RM0.00) - don't show that month at all rather
    # than a confusing zero payslip for a month they weren't employed.
    date_joined = payroll_calc._parse_date(emp["date_joined"])
    last_working_day = payroll_calc._parse_date(emp["last_working_day"])
    def _was_employed(year, month):
        days_in_month = calendar.monthrange(year, month)[1]
        month_start = datetime.date(year, month, 1)
        month_end = datetime.date(year, month, days_in_month)
        if date_joined and date_joined > month_end:
            return False
        if last_working_day and last_working_day < month_start:
            return False
        return True
    rows = [r for r in rows if _was_employed(r["year"], r["month"])]
    return render_template("portal_payslips.html", emp=emp, rows=rows)


@app.route("/portal/payslips/<int:year>/<int:month>")
@portal_login_required
def portal_payslip(year, month):
    db = get_db()
    emp = current_portal_employee(db)
    finalized = db.execute(
        "SELECT 1 FROM payroll_runs WHERE emp_id=? AND year=? AND month=?",
        (emp["emp_id"], year, month),
    ).fetchone()
    if not finalized:
        return "Payslip not available for this month yet.", 404
    if datetime.date.today() < payslip_release_date(db, year, month):
        return "This payslip hasn't been released yet. Please check back later.", 404
    result = payroll_calc.get_payroll_result(db, emp["emp_id"], year, month)
    return render_template("portal_payslip.html", emp=emp, r=result, year=year, month=month,
                            payment_date=payment_date_for(year, month))


@app.route("/portal/attendance")
@portal_login_required
def portal_attendance():
    db = get_db()
    emp = current_portal_employee(db)
    year = request.args.get("year", type=int) or datetime.date.today().year

    rows = db.execute(
        """SELECT * FROM attendance_monthly WHERE emp_id=? AND year=? ORDER BY month""",
        (emp["emp_id"], year),
    ).fetchall()

    # A row can exist for a month before someone actually joined (or after
    # they left) from a bulk-seeded/imported attendance table - exclude
    # those from what the employee sees and from their leave-balance sums.
    date_joined = payroll_calc._parse_date(emp["date_joined"])
    last_working_day = payroll_calc._parse_date(emp["last_working_day"])
    def _was_employed(month):
        days_in_month = calendar.monthrange(year, month)[1]
        month_start = datetime.date(year, month, 1)
        month_end = datetime.date(year, month, days_in_month)
        if date_joined and date_joined > month_end:
            return False
        if last_working_day and last_working_day < month_start:
            return False
        return True
    today = datetime.date.today()
    def _not_future(month):
        return (year, month) <= (today.year, today.month)
    rows = [r for r in rows if _was_employed(r["month"]) and _not_future(r["month"])]

    al_used = sum(r["al_days"] or 0 for r in rows)
    mc_used = sum(r["mc_days"] or 0 for r in rows)
    hl_used = sum(r["hl_days"] or 0 for r in rows)
    prorated_al, _al_note = _prorated_al_note(emp, year)
    al_entitlement = prorated_al if prorated_al is not None else (emp["annual_leave_entitlement"] or 0)
    al_balance = al_entitlement - al_used
    mc_balance = (emp["mc_entitlement"] or 0) - mc_used
    hl_balance = (emp["hospitalisation_leave_entitlement"] or 0) - hl_used

    years = db.execute(
        "SELECT DISTINCT year FROM attendance_monthly WHERE emp_id=? ORDER BY year DESC",
        (emp["emp_id"],),
    ).fetchall()
    if not years:
        years = [{"year": year}]

    return render_template(
        "portal_attendance.html", emp=emp, rows=rows, year=year, years=years,
        al_used=al_used, mc_used=mc_used, al_balance=al_balance, mc_balance=mc_balance,
        hl_used=hl_used, hl_balance=hl_balance,
    )


@app.route("/portal/leave", methods=["GET", "POST"])
@portal_login_required
def portal_leave():
    db = get_db()
    emp = current_portal_employee(db)
    error = None
    DOC_REQUIRED_TYPES = {"Medical Leave", "Hospitalisation Leave"}
    if request.method == "POST":
        leave_type = request.form.get("leave_type", "").strip()
        start_date = request.form.get("start_date", "")
        end_date = request.form.get("end_date", "")
        reason = request.form.get("reason", "").strip() or None
        file = request.files.get("supporting_doc")
        has_file = file is not None and file.filename != ""
        if not leave_type or not start_date or not end_date:
            error = "Leave type, start date, and end date are required."
        elif end_date < start_date:
            error = "End date cannot be before start date."
        elif leave_type in DOC_REQUIRED_TYPES and not has_file:
            error = f"{leave_type} requires a supporting document (e.g. medical certificate) to be uploaded."
        elif has_file:
            original_name = secure_filename(file.filename)
            ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""
            if ext not in ALLOWED_DOC_EXTENSIONS:
                error = "Supporting document must be a PDF, Word file, or an image (JPG/PNG)."
        if error is None:
            days = (datetime.date.fromisoformat(end_date) - datetime.date.fromisoformat(start_date)).days + 1
            stored_name = None
            if has_file:
                emp_dir = os.path.join(UPLOAD_DIR, emp["emp_id"])
                os.makedirs(emp_dir, exist_ok=True)
                stored_name = f"{uuid.uuid4().hex}_{original_name}"
                file.save(os.path.join(emp_dir, stored_name))
            db.execute(
                """INSERT INTO leave_requests (emp_id, leave_type, start_date, end_date, days, reason, status,
                       submitted_at, supporting_doc_original, supporting_doc_stored)
                   VALUES (?,?,?,?,?,?,'Pending',?,?,?)""",
                (emp["emp_id"], leave_type, start_date, end_date, days, reason,
                 datetime.datetime.now().isoformat(timespec="seconds"),
                 original_name if has_file else None, stored_name),
            )
            db.commit()
            return redirect(url_for("portal_leave"))

    my_requests = db.execute(
        """SELECT * FROM leave_requests WHERE emp_id=? ORDER BY submitted_at DESC""",
        (emp["emp_id"],),
    ).fetchall()

    cur_year = datetime.date.today().year
    year_rows = db.execute(
        "SELECT * FROM attendance_monthly WHERE emp_id=? AND year=?", (emp["emp_id"], cur_year)
    ).fetchall()
    al_used = sum(r["al_days"] or 0 for r in year_rows)
    mc_used = sum(r["mc_days"] or 0 for r in year_rows)
    hl_used = sum(r["hl_days"] or 0 for r in year_rows)
    prorated_al, _al_note = _prorated_al_note(emp, cur_year)
    al_entitlement = prorated_al if prorated_al is not None else (emp["annual_leave_entitlement"] or 0)
    al_balance = al_entitlement - al_used
    mc_balance = (emp["mc_entitlement"] or 0) - mc_used
    hl_balance = (emp["hospitalisation_leave_entitlement"] or 0) - hl_used

    return render_template("portal_leave.html", emp=emp, requests=my_requests, error=error,
                            al_balance=al_balance, mc_balance=mc_balance, hl_balance=hl_balance)


@app.route("/portal/leave/<int:request_id>/document")
@portal_login_required
def portal_leave_document(request_id):
    db = get_db()
    emp = current_portal_employee(db)
    lr = db.execute(
        "SELECT * FROM leave_requests WHERE id=? AND emp_id=?", (request_id, emp["emp_id"])
    ).fetchone()
    if lr is None or not lr["supporting_doc_stored"]:
        abort(404)
    return send_from_directory(
        os.path.join(UPLOAD_DIR, emp["emp_id"]), lr["supporting_doc_stored"],
        as_attachment=False, download_name=lr["supporting_doc_original"],
    )


@app.route("/portal/medical-claim", methods=["GET", "POST"])
@portal_login_required
def portal_medical_claim():
    db = get_db()
    emp = current_portal_employee(db)
    error = None
    if request.method == "POST":
        claim_date = request.form.get("claim_date", "")
        amount_raw = request.form.get("amount", "")
        clinic_name = request.form.get("clinic_name", "").strip() or None
        description = request.form.get("description", "").strip() or None
        file = request.files.get("supporting_doc")
        has_file = file is not None and file.filename != ""
        try:
            amount = float(amount_raw)
        except ValueError:
            amount = None
        if not claim_date or amount is None or amount <= 0:
            error = "Date of treatment and a valid claim amount are required."
        elif not has_file:
            error = "Please attach the receipt/invoice for this claim."
        else:
            original_name = secure_filename(file.filename)
            ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""
            if ext not in ALLOWED_DOC_EXTENSIONS:
                error = "Receipt must be a PDF, Word file, or an image (JPG/PNG)."
        if error is None:
            emp_dir = os.path.join(UPLOAD_DIR, emp["emp_id"])
            os.makedirs(emp_dir, exist_ok=True)
            stored_name = f"{uuid.uuid4().hex}_{original_name}"
            file.save(os.path.join(emp_dir, stored_name))
            db.execute(
                """INSERT INTO medical_claims (emp_id, claim_date, amount, clinic_name, description, status,
                       submitted_at, supporting_doc_original, supporting_doc_stored)
                   VALUES (?,?,?,?,?,'Pending',?,?,?)""",
                (emp["emp_id"], claim_date, amount, clinic_name, description,
                 datetime.datetime.now().isoformat(timespec="seconds"),
                 original_name, stored_name),
            )
            db.commit()
            return redirect(url_for("portal_medical_claim"))

    my_claims = db.execute(
        """SELECT * FROM medical_claims WHERE emp_id=? ORDER BY submitted_at DESC""",
        (emp["emp_id"],),
    ).fetchall()

    cur_year = datetime.date.today().year
    claimed_this_year = db.execute(
        """SELECT COALESCE(SUM(amount),0) AS total FROM medical_claims
           WHERE emp_id=? AND status='Approved' AND claim_date LIKE ?""",
        (emp["emp_id"], f"{cur_year:04d}-%"),
    ).fetchone()["total"]
    claim_limit, _limit_note = _prorated_medical_claim_limit(emp, cur_year)
    claim_balance = claim_limit - claimed_this_year

    return render_template("portal_medical_claim.html", emp=emp, claims=my_claims, error=error,
                            claimed_this_year=claimed_this_year, claim_balance=claim_balance,
                            claim_limit=claim_limit)


@app.route("/portal/medical-claim/<int:claim_id>/document")
@portal_login_required
def portal_medical_claim_document(claim_id):
    db = get_db()
    emp = current_portal_employee(db)
    mc = db.execute(
        "SELECT * FROM medical_claims WHERE id=? AND emp_id=?", (claim_id, emp["emp_id"])
    ).fetchone()
    if mc is None or not mc["supporting_doc_stored"]:
        abort(404)
    return send_from_directory(
        os.path.join(UPLOAD_DIR, emp["emp_id"]), mc["supporting_doc_stored"],
        as_attachment=False, download_name=mc["supporting_doc_original"],
    )


@app.route("/portal/business-trip/<int:trip_id>/document")
@portal_login_required
def portal_business_trip_document(trip_id):
    db = get_db()
    emp = current_portal_employee(db)
    trip = db.execute(
        "SELECT * FROM business_trips WHERE id=? AND emp_id=?", (trip_id, emp["emp_id"])
    ).fetchone()
    if trip is None or not trip["supporting_doc_stored"]:
        abort(404)
    return send_from_directory(
        os.path.join(UPLOAD_DIR, emp["emp_id"]), trip["supporting_doc_stored"],
        as_attachment=False, download_name=trip["supporting_doc_original"],
    )


@app.route("/portal/business-trip", methods=["GET", "POST"])
@portal_login_required
def portal_business_trip():
    db = get_db()
    emp = current_portal_employee(db)
    error = None
    if request.method == "POST":
        notice_type = request.form.get("notice_type") or "Business Trip"
        if notice_type not in BUSINESS_TRIP_TYPES:
            notice_type = "Business Trip"
        destination = request.form.get("destination", "").strip()
        start_date = request.form.get("start_date", "")
        end_date = request.form.get("end_date", "")
        purpose = request.form.get("purpose", "").strip() or None
        file = request.files.get("supporting_doc")
        has_file = file is not None and file.filename != ""
        original_name = stored_name = None
        if not destination or not start_date or not end_date:
            error = "Destination, start date, and end date are required."
        elif end_date < start_date:
            error = "End date cannot be before start date."
        elif notice_type == "Unrecorded Leave" and not has_file:
            error = "Please attach a supporting document for Unrecorded Leave."
        elif has_file:
            original_name = secure_filename(file.filename)
            ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""
            if ext not in ALLOWED_DOC_EXTENSIONS:
                error = "Supporting document must be a PDF, Word file, or an image (JPG/PNG)."
        if error is None:
            if has_file:
                emp_dir = os.path.join(UPLOAD_DIR, emp["emp_id"])
                os.makedirs(emp_dir, exist_ok=True)
                stored_name = f"{uuid.uuid4().hex}_{original_name}"
                file.save(os.path.join(emp_dir, stored_name))
            db.execute(
                """INSERT INTO business_trips (
                       emp_id, notice_type, destination, start_date, end_date, purpose,
                       supporting_doc_original, supporting_doc_stored, status, submitted_at
                   ) VALUES (?,?,?,?,?,?,?,?,'Pending',?)""",
                (emp["emp_id"], notice_type, destination, start_date, end_date, purpose,
                 original_name, stored_name,
                 datetime.datetime.now().isoformat(timespec="seconds")),
            )
            db.commit()
            return redirect(url_for("portal_business_trip"))

    my_trips = db.execute(
        "SELECT * FROM business_trips WHERE emp_id=? ORDER BY submitted_at DESC", (emp["emp_id"],)
    ).fetchall()
    return render_template("portal_business_trip.html", emp=emp, trips=my_trips, error=error,
                            notice_types=BUSINESS_TRIP_TYPES)


@app.route("/portal/ot-claim", methods=["GET", "POST"])
@portal_login_required
def portal_ot_claim():
    db = get_db()
    emp = current_portal_employee(db)
    error = None
    if request.method == "POST":
        claim_date = request.form.get("claim_date", "")
        ot_1_5 = request.form.get("ot_hours_1_5", type=float) or 0
        ot_2_0 = request.form.get("ot_hours_2_0", type=float) or 0
        ot_3_0 = request.form.get("ot_hours_3_0", type=float) or 0
        reason = request.form.get("reason", "").strip() or None
        if not claim_date:
            error = "Date is required."
        elif ot_1_5 <= 0 and ot_2_0 <= 0 and ot_3_0 <= 0:
            error = "Enter at least one OT hours amount."
        else:
            db.execute(
                """INSERT INTO ot_claims (emp_id, claim_date, ot_hours_1_5, ot_hours_2_0, ot_hours_3_0,
                       reason, status, submitted_by, submitted_at)
                   VALUES (?,?,?,?,?,?,'Pending','Employee',?)""",
                (emp["emp_id"], claim_date, ot_1_5, ot_2_0, ot_3_0, reason,
                 datetime.datetime.now().isoformat(timespec="seconds")),
            )
            db.commit()
            return redirect(url_for("portal_ot_claim"))
    my_claims = db.execute(
        "SELECT * FROM ot_claims WHERE emp_id=? ORDER BY submitted_at DESC", (emp["emp_id"],)
    ).fetchall()
    return render_template("portal_ot_claim.html", emp=emp, claims=my_claims, error=error)


@app.route("/portal/photo")
@portal_login_required
def portal_photo():
    """Serves the logged-in employee's own photo. Always resolves the
    employee from the portal session (never a URL-supplied emp_id), so an
    employee can only ever fetch their own photo, not anyone else's."""
    db = get_db()
    emp = current_portal_employee(db)
    if not emp["photo_path"]:
        abort(404)
    return send_from_directory(os.path.join(UPLOAD_DIR, emp["emp_id"]), emp["photo_path"])


@app.route("/portal/photo/upload", methods=["POST"])
@portal_login_required
def portal_photo_upload():
    """Lets an employee upload their own profile photo from the Staff
    Portal. Same storage/validation as the HR-side upload_photo route, but
    always targets the logged-in employee's own record."""
    db = get_db()
    emp = current_portal_employee(db)
    file = request.files.get("photo")
    if file is None or file.filename == "":
        return redirect(url_for("portal_profile"))

    original_name = secure_filename(file.filename)
    ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""
    if ext not in ALLOWED_PHOTO_EXTENSIONS:
        return "File type not allowed. Use JPG or PNG.", 400

    emp_dir = os.path.join(UPLOAD_DIR, emp["emp_id"])
    os.makedirs(emp_dir, exist_ok=True)
    if emp["photo_path"]:
        old_path = os.path.join(emp_dir, emp["photo_path"])
        if os.path.exists(old_path):
            os.remove(old_path)

    stored_name = f"photo_{uuid.uuid4().hex}.{ext}"
    file.save(os.path.join(emp_dir, stored_name))
    db.execute("UPDATE employees SET photo_path=? WHERE emp_id=?", (stored_name, emp["emp_id"]))
    db.commit()
    return redirect(url_for("portal_profile"))


@app.route("/portal/ic-upload", methods=["POST"])
@portal_login_required
def portal_ic_upload():
    """Lets an employee submit a photo/scan of their IC or passport from
    the Staff Portal, stored the same way as an HR-uploaded document
    (employee_documents, doc_type='IC / Passport Copy') so it shows up in
    Employee Edit's Documents section too. Always targets the logged-in
    employee's own record, never a URL-supplied emp_id."""
    db = get_db()
    emp = current_portal_employee(db)
    file = request.files.get("file")
    if file is None or file.filename == "":
        return redirect(url_for("portal_profile"))

    original_name = secure_filename(file.filename)
    ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""
    if ext not in ALLOWED_DOC_EXTENSIONS:
        return "File type not allowed. Use PDF, Word, or an image (JPG/PNG).", 400

    emp_dir = os.path.join(UPLOAD_DIR, emp["emp_id"])
    os.makedirs(emp_dir, exist_ok=True)
    stored_name = f"{uuid.uuid4().hex}_{original_name}"
    file.save(os.path.join(emp_dir, stored_name))

    db.execute(
        """INSERT INTO employee_documents (emp_id, doc_type, original_name, stored_name, notes, uploaded_at)
           VALUES (?,?,?,?,?,?)""",
        (emp["emp_id"], "IC / Passport Copy", original_name, stored_name, None,
         datetime.datetime.now().isoformat(timespec="seconds")),
    )
    db.commit()
    return redirect(url_for("portal_profile"))


@app.route("/portal/documents/<int:doc_id>/download")
@portal_login_required
def portal_document_download(doc_id):
    db = get_db()
    emp = current_portal_employee(db)
    doc = db.execute(
        "SELECT * FROM employee_documents WHERE id=? AND emp_id=?", (doc_id, emp["emp_id"]),
    ).fetchone()
    if doc is None:
        abort(404)
    return send_from_directory(os.path.join(UPLOAD_DIR, emp["emp_id"]), doc["stored_name"])


@app.route("/portal/profile", methods=["GET", "POST"])
@portal_login_required
def portal_profile():
    db = get_db()
    emp = current_portal_employee(db)
    saved = False
    if request.method == "POST":
        fields = {
            "phone_number": request.form.get("phone_number") or None,
            "hp_no": request.form.get("hp_no") or None,
            "email": request.form.get("email") or None,
            "address": request.form.get("address") or None,
            "marital_status": request.form.get("marital_status") or None,
            "religion": request.form.get("religion") or None,
            "emergency_contact_1_name": request.form.get("emergency_contact_1_name") or None,
            "emergency_contact_1_phone": request.form.get("emergency_contact_1_phone") or None,
            "emergency_contact_1_relationship": request.form.get("emergency_contact_1_relationship") or None,
            "emergency_contact_2_name": request.form.get("emergency_contact_2_name") or None,
            "emergency_contact_2_phone": request.form.get("emergency_contact_2_phone") or None,
            "emergency_contact_2_relationship": request.form.get("emergency_contact_2_relationship") or None,
        }
        new_password = request.form.get("new_password", "").strip()
        if new_password:
            fields["portal_password_hash"] = generate_password_hash(new_password)
        set_clause = ",".join(f"{col}=?" for col in fields)
        db.execute(
            f"UPDATE employees SET {set_clause} WHERE emp_id=?",
            list(fields.values()) + [emp["emp_id"]],
        )
        db.execute(
            "INSERT INTO profile_update_log (emp_id, updated_at) VALUES (?,?)",
            (emp["emp_id"], datetime.datetime.now().isoformat(timespec="seconds")),
        )
        db.commit()
        emp = current_portal_employee(db)
        saved = True
    ic_documents = db.execute(
        """SELECT * FROM employee_documents WHERE emp_id=? AND doc_type='IC / Passport Copy'
           ORDER BY uploaded_at DESC""",
        (emp["emp_id"],),
    ).fetchall()
    return render_template("portal_profile.html", emp=emp, saved=saved, ic_documents=ic_documents,
                            religion_options=RELIGION_OPTIONS)


# ---------------- HR: Leave Taken Report ----------------

def _prorate_by_days(e, year, full):
    """Malaysian practice (EA1955) is to prorate a calendar-year entitlement
    (Annual Leave, medical claim limit, etc.) by days of service in that
    year - both for someone who resigned partway through the year, and for
    someone who joined partway through the year (whether they're still
    active or have since resigned). Prorated by calendar days (not
    completed months) so someone employed the whole year - e.g. joined 1
    Jan, still active - always comes out to the full amount. `full` is the
    already-resolved full-year amount to prorate. Returns
    (prorated_amount, note), or (None, note) if there isn't enough data on
    file (Date Joined / Last Working Day) to prorate reliably - in that
    case we refuse to guess rather than show a made-up number. Returns
    (None, None) if the person was employed for the whole of `year`, so no
    proration is needed at all.
    """
    dj = e["date_joined"]
    lwd = e["last_working_day"]
    joined_this_year = bool(dj) and dj.startswith(str(year))
    resigned_this_year = bool(lwd) and lwd.startswith(str(year))

    if not joined_this_year and not resigned_this_year:
        resignation_date = e["resignation_date"]
        if not lwd and resignation_date and resignation_date.startswith(str(year)):
            return None, "resignation on file ({}) but no Last Working Day - can't prorate".format(resignation_date)
        return None, None  # employed for the whole year - no proration needed
    if resigned_this_year and not dj:
        return None, "resigned {} - missing Date Joined, can't prorate".format(lwd)

    start_of_year = datetime.date(year, 1, 1)
    end_of_year = datetime.date(year, 12, 31)
    start = max(datetime.date.fromisoformat(dj), start_of_year) if dj else start_of_year
    end = datetime.date.fromisoformat(lwd) if resigned_this_year else end_of_year
    days_employed = max((end - start).days + 1, 0)
    days_in_year = (end_of_year - start_of_year).days + 1  # 365 or 366
    prorated = round((full or 0) * days_employed / days_in_year, 2)
    if resigned_this_year:
        note = "resigned {} - prorated for {} of {} days in {}".format(lwd, days_employed, days_in_year, year)
    else:
        note = "joined {} - prorated for {} of {} days in {}".format(dj, days_employed, days_in_year, year)
    return prorated, note


def _prorated_al_note(e, year):
    return _prorate_by_days(e, year, e["annual_leave_entitlement"])


def _prorated_medical_claim_limit(e, year):
    """Same day-based proration as Annual Leave, applied to the RM/year
    medical claim limit."""
    prorated, note = _prorate_by_days(e, year, e["medical_claim_limit"])
    return prorated if prorated is not None else (e["medical_claim_limit"] or 0), note


@app.route("/leave-report")
def leave_report():
    db = get_db()
    year = request.args.get("year", type=int) or datetime.date.today().year

    emps = db.execute(
        """SELECT emp_id, full_name, department, date_joined, last_working_day, resignation_date,
                  annual_leave_entitlement, mc_entitlement
           FROM employees
           WHERE status != 'Inactive'
              OR (last_working_day IS NOT NULL AND last_working_day LIKE ?)
              OR (resignation_date IS NOT NULL AND resignation_date LIKE ? AND last_working_day IS NULL)
           ORDER BY emp_id""",
        (f"{year}%", f"{year}%"),
    ).fetchall()

    rows = []
    for e in emps:
        totals = db.execute(
            """SELECT COALESCE(SUM(al_days),0) AS al, COALESCE(SUM(mc_days),0) AS mc,
                      COALESCE(SUM(hl_days),0) AS hl, COALESCE(SUM(ul_days),0) AS ul,
                      COALESCE(SUM(other_paid_leave),0) AS other
               FROM attendance_monthly WHERE emp_id=? AND year=?""",
            (e["emp_id"], year),
        ).fetchone()
        al_entitlement = e["annual_leave_entitlement"] or 0
        mc_entitlement = e["mc_entitlement"] or 0
        prorated_al, al_note = _prorated_al_note(e, year)
        if prorated_al is not None:
            al_entitlement = prorated_al
        rows.append({
            "emp_id": e["emp_id"], "full_name": e["full_name"], "department": e["department"],
            "al_entitlement": al_entitlement, "al_used": totals["al"],
            "al_balance": al_entitlement - totals["al"], "al_note": al_note,
            "mc_entitlement": mc_entitlement, "mc_used": totals["mc"],
            "mc_balance": mc_entitlement - totals["mc"],
            "hl_used": totals["hl"], "ul_used": totals["ul"], "other_used": totals["other"],
        })

    years = db.execute("SELECT DISTINCT year FROM attendance_monthly ORDER BY year DESC").fetchall()
    if not years:
        years = [{"year": year}]

    return render_template("leave_report.html", rows=rows, year=year, years=years)


# ---------------- HR: Leave Requests Admin ----------------

@app.route("/leave-requests")
def leave_requests_admin():
    db = get_db()
    # An approver (e.g. Mr Kee) only sees requests for employees assigned to
    # them; HR (role='admin') sees everyone - matches "all leave except
    # W001/W002/K002/K003 is under Mr Kee or HR".
    scope_clause = ""
    params = []
    if session.get("hr_role") == "approver":
        scope_clause = "AND e.leave_approver_username=?"
        params.append(session["hr_username"])
    pending = db.execute(
        f"""SELECT lr.*, e.full_name FROM leave_requests lr
           JOIN employees e ON e.emp_id = lr.emp_id
           WHERE lr.status='Pending' {scope_clause} ORDER BY lr.submitted_at""",
        params,
    ).fetchall()

    # Reviewed (historical) list is filtered to one year at a time - Jan to
    # Dec of whichever year is selected - by the leave's own start_date
    # rather than an arbitrary "last 50" cap.
    year = request.args.get("year", type=int) or datetime.date.today().year
    reviewed = db.execute(
        f"""SELECT lr.*, e.full_name FROM leave_requests lr
           JOIN employees e ON e.emp_id = lr.emp_id
           WHERE lr.status!='Pending' AND lr.start_date LIKE ? {scope_clause}
           ORDER BY lr.reviewed_at DESC""",
        [f"{year}%"] + params,
    ).fetchall()
    years = db.execute(
        f"""SELECT DISTINCT CAST(substr(lr.start_date,1,4) AS INTEGER) AS year
           FROM leave_requests lr JOIN employees e ON e.emp_id = lr.emp_id
           WHERE lr.status!='Pending' {scope_clause} ORDER BY year DESC""",
        params,
    ).fetchall()
    if not years or year not in [y["year"] for y in years]:
        years = list(years) + [{"year": year}]
        years.sort(key=lambda y: y["year"], reverse=True)

    return render_template("leave_requests_admin.html", pending=pending, reviewed=reviewed,
                            year=year, years=years)


@app.route("/leave-requests/<int:request_id>/delete", methods=["POST"])
def delete_leave_request(request_id):
    """HR-only: removes a leave request record entirely (e.g. test/bad
    data), rather than just changing its status. Does not touch
    attendance_monthly - HR keys leave days there separately (see the
    leave_requests table's own comment), so there's nothing else to undo."""
    if session.get("hr_role") != "admin":
        abort(403)
    db = get_db()
    db.execute("DELETE FROM leave_requests WHERE id=?", (request_id,))
    db.commit()
    return redirect(url_for("leave_requests_admin"))


@app.route("/leave-requests/<int:request_id>/document")
def leave_request_document(request_id):
    db = get_db()
    lr = db.execute(
        """SELECT lr.*, e.leave_approver_username FROM leave_requests lr
           JOIN employees e ON e.emp_id = lr.emp_id WHERE lr.id=?""",
        (request_id,),
    ).fetchone()
    if lr is None or not lr["supporting_doc_stored"]:
        abort(404)
    if session.get("hr_role") == "approver" and lr["leave_approver_username"] != session["hr_username"]:
        abort(403)
    return send_from_directory(
        os.path.join(UPLOAD_DIR, lr["emp_id"]), lr["supporting_doc_stored"],
        as_attachment=False, download_name=lr["supporting_doc_original"],
    )


LEAVE_TYPE_TO_ATTENDANCE_COLUMN = {
    "Annual Leave": "al_days",
    "Medical Leave": "mc_days",
    "Hospitalisation Leave": "hl_days",
    "Unpaid Leave": "ul_days",
    # No dedicated columns for these two - attendance_monthly.other_paid_leave
    # is documented as "PL + EL" (Paid Leave + Emergency Leave), which covers
    # both.
    "Maternity/Paternity Leave": "other_paid_leave",
    "Emergency Leave": "other_paid_leave",
}


def _sync_leave_to_attendance(db, leave_request):
    """Adds an approved leave request's days onto Attendance, split across
    whichever month(s) the date range actually falls in, so HR doesn't have
    to re-enter the same days by hand on the Attendance page. Adds to
    whatever's already there for that month rather than overwriting it,
    since other days may already be recorded."""
    column = LEAVE_TYPE_TO_ATTENDANCE_COLUMN.get(leave_request["leave_type"])
    if column is None:
        return  # unrecognized leave type - nothing to sync

    start = datetime.date.fromisoformat(leave_request["start_date"])
    end = datetime.date.fromisoformat(leave_request["end_date"])

    days_per_month = {}
    day = start
    while day <= end:
        key = (day.year, day.month)
        days_per_month[key] = days_per_month.get(key, 0) + 1
        day += datetime.timedelta(days=1)

    for (year, month), count in days_per_month.items():
        db.execute(
            f"""INSERT INTO attendance_monthly (emp_id, year, month, {column})
                VALUES (?,?,?,?)
                ON CONFLICT(emp_id, year, month) DO UPDATE SET
                    {column} = {column} + excluded.{column}""",
            (leave_request["emp_id"], year, month, count),
        )
        # Days Worked and Meal Eligible Days both need to drop when leave is
        # added, or the employee ends up credited for days off as if they
        # were fully worked (and meal-allowance-eligible). Recompute both
        # from Working Days in Month minus every leave/absence type on file,
        # rather than just decrementing, so they stay correct even if this
        # runs more than once or attendance was edited independently.
        row = db.execute(
            """SELECT working_days_in_month, al_days, mc_days, hl_days,
                      ul_days, other_paid_leave, absent_days
               FROM attendance_monthly WHERE emp_id=? AND year=? AND month=?""",
            (leave_request["emp_id"], year, month),
        ).fetchone()
        total_leave = sum(
            row[k] or 0 for k in
            ("al_days", "mc_days", "hl_days", "ul_days", "other_paid_leave", "absent_days")
        )
        recomputed_days = max((row["working_days_in_month"] or 0) - total_leave, 0)
        db.execute(
            """UPDATE attendance_monthly SET days_worked=?, meal_eligible_days=?
               WHERE emp_id=? AND year=? AND month=?""",
            (recomputed_days, recomputed_days, leave_request["emp_id"], year, month),
        )


@app.route("/leave-requests/<int:request_id>/review", methods=["POST"])
def review_leave_request(request_id):
    db = get_db()
    decision = request.form.get("decision")
    if decision not in ("Approved", "Rejected"):
        return "Invalid decision", 400
    notes = request.form.get("review_notes") or None
    leave_request = db.execute(
        """SELECT lr.*, e.leave_approver_username FROM leave_requests lr
           JOIN employees e ON e.emp_id = lr.emp_id WHERE lr.id=?""",
        (request_id,),
    ).fetchone()
    if leave_request is None:
        return "Leave request not found", 404
    if session.get("hr_role") == "approver" and leave_request["leave_approver_username"] != session["hr_username"]:
        abort(403)
    # Reviewer is derived from who's actually logged in, not typed by hand -
    # ties every decision to a real account now that HR/approver logins exist.
    hr_user = db.execute("SELECT full_name FROM hr_users WHERE username=?", (session["hr_username"],)).fetchone()
    reviewer = hr_user["full_name"] if hr_user else session["hr_username"]
    db.execute(
        """UPDATE leave_requests SET status=?, reviewed_by=?, reviewed_at=?, review_notes=?
           WHERE id=?""",
        (decision, reviewer, datetime.datetime.now().isoformat(timespec="seconds"), notes, request_id),
    )
    if decision == "Approved":
        _sync_leave_to_attendance(db, leave_request)
    db.commit()
    return redirect(url_for("leave_requests_admin"))


# ---------------- HR: Business Trips Admin ----------------

@app.route("/business-trips")
def business_trips_admin():
    db = get_db()
    # An approver (e.g. Mr Kee) only sees notices for employees assigned to
    # them as Leave Approver - reuses that same assignment rather than a
    # separate one, since it's the same "who's this person's supervisor"
    # relationship. HR (role='admin') sees everyone.
    scope_clause = ""
    params = []
    if session.get("hr_role") == "approver":
        scope_clause = "AND e.leave_approver_username=?"
        params.append(session["hr_username"])
    pending = db.execute(
        f"""SELECT bt.*, e.full_name FROM business_trips bt
           JOIN employees e ON e.emp_id = bt.emp_id
           WHERE bt.status='Pending' {scope_clause} ORDER BY bt.submitted_at""",
        params,
    ).fetchall()
    reviewed = db.execute(
        f"""SELECT bt.*, e.full_name FROM business_trips bt
           JOIN employees e ON e.emp_id = bt.emp_id
           WHERE bt.status!='Pending' {scope_clause} ORDER BY bt.reviewed_at DESC LIMIT 50""",
        params,
    ).fetchall()
    return render_template("business_trips_admin.html", pending=pending, reviewed=reviewed)


@app.route("/business-trips/<int:trip_id>/document")
def business_trip_document(trip_id):
    db = get_db()
    trip = db.execute("SELECT * FROM business_trips WHERE id=?", (trip_id,)).fetchone()
    if trip is None or not trip["supporting_doc_stored"]:
        abort(404)
    return send_from_directory(
        os.path.join(UPLOAD_DIR, trip["emp_id"]), trip["supporting_doc_stored"],
        as_attachment=False, download_name=trip["supporting_doc_original"],
    )


@app.route("/business-trips/<int:trip_id>/review", methods=["POST"])
def review_business_trip(trip_id):
    db = get_db()
    decision = request.form.get("decision")
    if decision not in ("Approved", "Rejected"):
        return "Invalid decision", 400
    notes = request.form.get("review_notes") or None
    trip = db.execute(
        """SELECT bt.*, e.leave_approver_username FROM business_trips bt
           JOIN employees e ON e.emp_id = bt.emp_id WHERE bt.id=?""",
        (trip_id,),
    ).fetchone()
    if trip is None:
        return "Business trip not found", 404
    if session.get("hr_role") == "approver" and trip["leave_approver_username"] != session["hr_username"]:
        abort(403)
    hr_user = db.execute("SELECT full_name FROM hr_users WHERE username=?", (session["hr_username"],)).fetchone()
    reviewer = hr_user["full_name"] if hr_user else session["hr_username"]
    db.execute(
        """UPDATE business_trips SET status=?, reviewed_by=?, reviewed_at=?, review_notes=?
           WHERE id=?""",
        (decision, reviewer, datetime.datetime.now().isoformat(timespec="seconds"), notes, trip_id),
    )
    db.commit()
    return redirect(url_for("business_trips_admin"))


# ---------------- HR: OT Claims Admin ----------------
# For employees.ot_approval_required='Y' (e.g. executives who normally
# aren't OT-eligible but can claim OT with Director approval) - unlike
# Leave Approver/Business Trips, this has a single company-wide approver
# (Mr Yang Hui, hr_users.can_approve_ot='Y'), not a per-employee assignment.

def _apply_ot_claim_to_attendance(db, emp_id, claim_date, ot_1_5, ot_2_0, ot_3_0):
    """Adds an approved OT claim's hours into that date's attendance_daily
    row (creating one as WORKED if it doesn't exist yet), then recomputes
    that month's attendance_monthly aggregate - the same path Daily
    Attendance itself feeds payroll through, so an approved claim counts
    exactly like HR having typed those hours in directly."""
    db.execute(
        """INSERT INTO attendance_daily (emp_id, date, ot_hours_1_5, ot_hours_2_0, ot_hours_3_0)
           VALUES (?,?,?,?,?)
           ON CONFLICT(emp_id, date) DO UPDATE SET
             ot_hours_1_5 = COALESCE(ot_hours_1_5,0) + excluded.ot_hours_1_5,
             ot_hours_2_0 = COALESCE(ot_hours_2_0,0) + excluded.ot_hours_2_0,
             ot_hours_3_0 = COALESCE(ot_hours_3_0,0) + excluded.ot_hours_3_0""",
        (emp_id, claim_date, ot_1_5, ot_2_0, ot_3_0),
    )
    year, month, _day = (int(p) for p in claim_date.split("-"))
    _sync_daily_to_monthly(db, emp_id, year, month)


@app.route("/ot-claims")
def ot_claims_admin():
    db = get_db()
    pending = db.execute(
        """SELECT oc.*, e.full_name FROM ot_claims oc
           JOIN employees e ON e.emp_id = oc.emp_id
           WHERE oc.status='Pending' ORDER BY oc.submitted_at"""
    ).fetchall()
    reviewed = db.execute(
        """SELECT oc.*, e.full_name FROM ot_claims oc
           JOIN employees e ON e.emp_id = oc.emp_id
           WHERE oc.status!='Pending' ORDER BY oc.reviewed_at DESC LIMIT 50"""
    ).fetchall()
    flagged_employees = db.execute(
        """SELECT emp_id, full_name FROM employees
           WHERE ot_approval_required='Y' AND (status IS NULL OR status != 'Inactive')
           ORDER BY full_name"""
    ).fetchall()
    approvers = db.execute(
        "SELECT full_name FROM hr_users WHERE can_approve_ot='Y' ORDER BY full_name"
    ).fetchall()
    approver_names = ", ".join(r["full_name"] for r in approvers) or "no one yet - see Settings"
    return render_template("ot_claims_admin.html", pending=pending, reviewed=reviewed,
                            flagged_employees=flagged_employees, approver_names=approver_names)


@app.route("/ot-claims/new", methods=["POST"])
def ot_claim_new():
    """HR keying in an OT claim from a paper attendance sheet on an
    employee's behalf - same Pending status and approval flow as one the
    employee submits themselves via the Staff Portal."""
    db = get_db()
    emp_id = request.form.get("emp_id")
    claim_date = request.form.get("claim_date", "")
    if db.execute("SELECT 1 FROM employees WHERE emp_id=?", (emp_id,)).fetchone() is None or not claim_date:
        return "Employee and claim date are required", 400
    ot_1_5 = request.form.get("ot_hours_1_5", type=float) or 0
    ot_2_0 = request.form.get("ot_hours_2_0", type=float) or 0
    ot_3_0 = request.form.get("ot_hours_3_0", type=float) or 0
    reason = request.form.get("reason", "").strip() or None
    db.execute(
        """INSERT INTO ot_claims (emp_id, claim_date, ot_hours_1_5, ot_hours_2_0, ot_hours_3_0,
               reason, status, submitted_by, submitted_at)
           VALUES (?,?,?,?,?,?,'Pending',?,?)""",
        (emp_id, claim_date, ot_1_5, ot_2_0, ot_3_0, reason,
         session["hr_username"], datetime.datetime.now().isoformat(timespec="seconds")),
    )
    db.commit()
    return redirect(url_for("ot_claims_admin"))


@app.route("/ot-claims/<int:claim_id>/review", methods=["POST"])
def review_ot_claim(claim_id):
    db = get_db()
    decision = request.form.get("decision")
    if decision not in ("Approved", "Rejected"):
        return "Invalid decision", 400
    notes = request.form.get("review_notes") or None
    claim = db.execute("SELECT * FROM ot_claims WHERE id=?", (claim_id,)).fetchone()
    if claim is None:
        return "OT claim not found", 404
    hr_user = db.execute("SELECT full_name FROM hr_users WHERE username=?", (session["hr_username"],)).fetchone()
    reviewer = hr_user["full_name"] if hr_user else session["hr_username"]
    db.execute(
        """UPDATE ot_claims SET status=?, reviewed_by=?, reviewed_at=?, review_notes=? WHERE id=?""",
        (decision, reviewer, datetime.datetime.now().isoformat(timespec="seconds"), notes, claim_id),
    )
    if decision == "Approved":
        _apply_ot_claim_to_attendance(
            db, claim["emp_id"], claim["claim_date"],
            claim["ot_hours_1_5"], claim["ot_hours_2_0"], claim["ot_hours_3_0"],
        )
    db.commit()
    return redirect(url_for("ot_claims_admin"))


@app.route("/ot-claims/<int:claim_id>/delete", methods=["POST"])
def delete_ot_claim(claim_id):
    """Permanently removes an OT claim record - refused for an Approved
    claim, since its hours are already reflected in attendance/payroll
    and deleting the record would leave that with no audit trail."""
    db = get_db()
    claim = db.execute("SELECT status FROM ot_claims WHERE id=?", (claim_id,)).fetchone()
    if claim is None:
        return "OT claim not found", 404
    if claim["status"] == "Approved":
        return "Refused: cannot delete an Approved claim - its hours are already reflected in attendance/payroll", 400
    db.execute("DELETE FROM ot_claims WHERE id=?", (claim_id,))
    db.commit()
    return redirect(url_for("ot_claims_admin"))


# ---------------- HR: Medical Claims Admin ----------------

@app.route("/medical-claims")
def medical_claims_admin():
    db = get_db()
    pending = db.execute(
        """SELECT mc.*, e.full_name FROM medical_claims mc
           JOIN employees e ON e.emp_id = mc.emp_id
           WHERE mc.status='Pending' ORDER BY mc.submitted_at"""
    ).fetchall()
    reviewed = db.execute(
        """SELECT mc.*, e.full_name FROM medical_claims mc
           JOIN employees e ON e.emp_id = mc.emp_id
           WHERE mc.status!='Pending' ORDER BY mc.reviewed_at DESC LIMIT 50"""
    ).fetchall()
    return render_template("medical_claims_admin.html", pending=pending, reviewed=reviewed)


@app.route("/medical-claims/<int:claim_id>/document")
def medical_claim_document(claim_id):
    db = get_db()
    mc = db.execute("SELECT * FROM medical_claims WHERE id=?", (claim_id,)).fetchone()
    if mc is None or not mc["supporting_doc_stored"]:
        abort(404)
    return send_from_directory(
        os.path.join(UPLOAD_DIR, mc["emp_id"]), mc["supporting_doc_stored"],
        as_attachment=False, download_name=mc["supporting_doc_original"],
    )


@app.route("/medical-claims/<int:claim_id>/review", methods=["POST"])
def review_medical_claim(claim_id):
    db = get_db()
    decision = request.form.get("decision")
    if decision not in ("Approved", "Rejected"):
        return "Invalid decision", 400
    notes = request.form.get("review_notes") or None
    claim = db.execute("SELECT * FROM medical_claims WHERE id=?", (claim_id,)).fetchone()
    if claim is None:
        return "Medical claim not found", 404
    hr_user = db.execute("SELECT full_name FROM hr_users WHERE username=?", (session["hr_username"],)).fetchone()
    reviewer = hr_user["full_name"] if hr_user else session["hr_username"]
    db.execute(
        """UPDATE medical_claims SET status=?, reviewed_by=?, reviewed_at=?, review_notes=?
           WHERE id=?""",
        (decision, reviewer, datetime.datetime.now().isoformat(timespec="seconds"), notes, claim_id),
    )
    db.commit()
    return redirect(url_for("medical_claims_admin"))


# ---------------- Appraisals ----------------
# Digitized version of the paper Performance Appraisal Form
# (confirmation_appraisal.html) - same rating factors, 1-5 scale, and
# overall evaluation bands. Scope for now: the supervisor's rating +
# recommendation only, not the paper form's later Manager/Director
# countersign, HR verification, or salary sections.

APPRAISAL_CATEGORIES = [
    ("Behavior", ["Attitude", "Attendance", "Punctuality", "Discipline", "Integrity/Honesty"]),
    ("Job Performance", ["Commitment", "Cooperation & Team Work", "Dependability & Adaptability",
                          "Initiative", "Potentiality", "Quality Of Work"]),
    ("Knowledge", ["House Keeping (5S)", "Safety, Health & Environment"]),
    ("Skills", ["Communication", "Decision Making", "Planning & Organizing Ability",
                "Problem Solving", "Supervision / Leadership", "Time Management"]),
]
APPRAISAL_FACTORS = [f for _, factors in APPRAISAL_CATEGORIES for f in factors]
APPRAISAL_PURPOSES = ["Confirmation", "Promotion", "Increment", "Bonus", "Assessment"]
# Matches confirmation_appraisal.html's band table exactly (percentage of max score).
APPRAISAL_BANDS = [(86, "Excellent"), (66, "Good"), (51, "Average"), (36, "Fair"), (0, "Poor")]


def _appraisal_band(percentage):
    for threshold, label in APPRAISAL_BANDS:
        if percentage >= threshold:
            return label
    return "Poor"


def _require_appraisal_access():
    """role='admin' always has access; role='approver' needs can_approve_appraisal='Y'
    (already enforced by the before_request gate for the /appraisals prefix, but
    individual routes still need to know which supervisor's team to scope to)."""
    if session.get("hr_role") == "admin":
        return None  # unscoped - sees everyone
    return session.get("hr_username")  # scoped to this supervisor's assigned staff


@app.route("/appraisals/team")
def appraisal_team():
    """A supervisor's (Mr Kee's or Mr Yang's) list of assigned staff, each
    with their most recent appraisal status. Admin visiting this directly
    sees everyone (no supervisor assignment required)."""
    db = get_db()
    supervisor_username = _require_appraisal_access()
    # Already-confirmed staff have nothing left to do here (this list is
    # about who's still pending confirmation) so they're excluded entirely,
    # not just pushed down. Sorted by Confirm Due Date ascending (soonest
    # due at the top) so it reads like a to-do list; anyone with no
    # probation_end_date on file sorts last.
    confirmed_clause = "AND (confirmation_date IS NULL OR confirmation_date = '')"
    order_by = "ORDER BY (probation_end_date IS NULL OR probation_end_date = ''), probation_end_date, emp_id"
    if supervisor_username:
        staff = db.execute(
            f"""SELECT emp_id, full_name, position, date_joined, probation_end_date
               FROM employees WHERE appraisal_supervisor_username=? {confirmed_clause} {order_by}""",
            (supervisor_username,),
        ).fetchall()
    else:
        staff = db.execute(
            f"""SELECT emp_id, full_name, position, date_joined, probation_end_date
               FROM employees WHERE status != 'Inactive' {confirmed_clause} {order_by}"""
        ).fetchall()

    latest_by_emp = {}
    for row in db.execute(
        "SELECT emp_id, id, status, overall_band, appraisal_date FROM appraisals ORDER BY id DESC"
    ).fetchall():
        latest_by_emp.setdefault(row["emp_id"], row)

    return render_template("appraisal_team.html", staff=staff, latest_by_emp=latest_by_emp,
                            today=datetime.date.today().isoformat())


@app.route("/appraisals")
def appraisals_admin():
    """HR's outcomes view - every submitted appraisal, company-wide.
    Admin-only in practice (an approver's /appraisals requests never reach
    here - the before_request gate only allows them onto /appraisals/team,
    /appraisals/<id> for their own submissions, and /appraisals/<emp_id>/new)."""
    db = get_db()
    rows = db.execute(
        """SELECT a.*, e.full_name FROM appraisals a
           JOIN employees e ON e.emp_id = a.emp_id
           WHERE a.status='Submitted' ORDER BY a.submitted_at DESC"""
    ).fetchall()
    return render_template("appraisals_admin.html", rows=rows)


@app.route("/appraisals/<emp_id>/new", methods=["GET", "POST"])
def appraisal_new(emp_id):
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE emp_id=?", (emp_id,)).fetchone()
    if emp is None:
        return "Employee not found", 404

    supervisor_username = _require_appraisal_access()
    if supervisor_username and emp["appraisal_supervisor_username"] != supervisor_username:
        abort(403)

    error = None
    if request.method == "POST":
        ratings = {}
        for factor in APPRAISAL_FACTORS:
            raw = request.form.get(f"rating__{factor}")
            try:
                score = int(raw)
            except (TypeError, ValueError):
                score = None
            if score is None or not 1 <= score <= 5:
                error = f"Please give a 1-5 rating for every factor (missing: {factor})."
                break
            ratings[factor] = score

        if error is None:
            total = sum(ratings.values())
            max_score = len(APPRAISAL_FACTORS) * 5
            percentage = round(total / max_score * 100, 1)
            current_salary = emp["basic_salary"] or 0
            try:
                increment_amount = float(request.form.get("increment_amount") or 0)
            except ValueError:
                increment_amount = 0
            new_salary = round(current_salary + increment_amount, 2)
            db.execute(
                """INSERT INTO appraisals (
                       emp_id, purpose, appraisal_date, ratings_json, total_score, max_score,
                       percentage, overall_band, rec_advancement, rec_not_yet_ready,
                       rec_better_suited, rec_training_required, comments,
                       current_salary, increment_amount, new_salary, status,
                       supervisor_username, submitted_at
                   ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'Submitted',?,?)""",
                (
                    emp_id, request.form.get("purpose", "Assessment"),
                    request.form.get("appraisal_date") or datetime.date.today().isoformat(),
                    json.dumps(ratings), total, max_score, percentage, _appraisal_band(percentage),
                    "Y" if request.form.get("rec_advancement") else "N",
                    "Y" if request.form.get("rec_not_yet_ready") else "N",
                    "Y" if request.form.get("rec_better_suited") else "N",
                    "Y" if request.form.get("rec_training_required") else "N",
                    request.form.get("comments") or None,
                    current_salary, increment_amount, new_salary,
                    session["hr_username"],
                    datetime.datetime.now().isoformat(timespec="seconds"),
                ),
            )
            db.commit()
            return redirect(url_for("appraisal_team"))

    return render_template(
        "appraisal_form.html", emp=emp, error=error, categories=APPRAISAL_CATEGORIES,
        purposes=APPRAISAL_PURPOSES, today=datetime.date.today(),
    )


@app.route("/appraisals/<int:appraisal_id>")
def appraisal_view(appraisal_id):
    db = get_db()
    row = db.execute(
        """SELECT a.*, e.full_name, e.position, e.department FROM appraisals a
           JOIN employees e ON e.emp_id = a.emp_id WHERE a.id=?""",
        (appraisal_id,),
    ).fetchone()
    if row is None:
        abort(404)
    supervisor_username = _require_appraisal_access()
    if supervisor_username and row["supervisor_username"] != supervisor_username:
        abort(403)
    if not supervisor_username and row["status"] == "Submitted" and not row["hr_viewed_at"]:
        # An HR admin (not the submitting approver) just looked at this -
        # clears it from the "new appraisal" nav badge.
        db.execute("UPDATE appraisals SET hr_viewed_at=? WHERE id=?",
                   (datetime.datetime.now().isoformat(timespec="seconds"), appraisal_id))
        db.commit()
        row = db.execute(
            """SELECT a.*, e.full_name, e.position, e.department FROM appraisals a
               JOIN employees e ON e.emp_id = a.emp_id WHERE a.id=?""",
            (appraisal_id,),
        ).fetchone()
    ratings = json.loads(row["ratings_json"])
    return render_template(
        "appraisal_view.html", a=row, ratings=ratings, categories=APPRAISAL_CATEGORIES,
    )


# ---------------- One-time data migration (local -> production) ----------------
# Lets the real payroll.db / uploads folder be copied from a local machine
# onto a fresh production deployment (Railway/Render/etc.) where they
# deliberately don't exist yet (never committed to GitHub). Gated by
# RESTORE_TOKEN, an env var only the deployer sets/knows - both routes
# fail closed (404) if it isn't configured, so they're inert everywhere
# except a deployment that's deliberately been put into "accepting a
# restore" mode. Remove RESTORE_TOKEN from the environment (or delete
# these two routes) once the migration is done.

@app.route("/hr/restore-database", methods=["POST"])
def hr_restore_database():
    token = os.environ.get("RESTORE_TOKEN")
    if not token or request.form.get("token") != token:
        abort(404)
    file = request.files.get("file")
    if file is None or file.filename == "":
        return "No file uploaded", 400
    if os.path.exists(DB_PATH):
        os.replace(DB_PATH, DB_PATH + ".bak")
    file.save(DB_PATH)
    return f"OK - database restored to {DB_PATH} (previous copy saved as payroll.db.bak)", 200


@app.route("/hr/restore-uploads", methods=["POST"])
def hr_restore_uploads():
    token = os.environ.get("RESTORE_TOKEN")
    if not token or request.form.get("token") != token:
        abort(404)
    file = request.files.get("file")
    if file is None or file.filename == "":
        return "No file uploaded", 400
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    zip_path = os.path.join(DATA_DIR, "_uploads_restore.zip")
    file.save(zip_path)
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(UPLOAD_DIR)
    os.remove(zip_path)
    return f"OK - uploads restored to {UPLOAD_DIR}", 200


@app.route("/hr/migrate-schema", methods=["POST"])
def hr_migrate_schema():
    """Additively applies schema changes that shipped after the database
    was created/restored - new columns (ALTER TABLE ... ADD COLUMN, skipped
    if already present) and new tables (CREATE TABLE IF NOT EXISTS) - never
    touches existing data. Same RESTORE_TOKEN gate as the restore routes
    above; safe to call more than once (every step checks first)."""
    token = os.environ.get("RESTORE_TOKEN")
    if not token or request.form.get("token") != token:
        abort(404)
    db = get_db()
    applied = []

    hr_cols = [r[1] for r in db.execute("PRAGMA table_info(hr_users)").fetchall()]
    for col, decl in [("can_approve_leave", "TEXT NOT NULL DEFAULT 'N'"),
                       ("can_approve_appraisal", "TEXT NOT NULL DEFAULT 'N'"),
                       ("can_approve_ot", "TEXT NOT NULL DEFAULT 'N'")]:
        if col not in hr_cols:
            db.execute(f"ALTER TABLE hr_users ADD COLUMN {col} {decl}")
            applied.append(f"hr_users.{col}")

    emp_cols = [r[1] for r in db.execute("PRAGMA table_info(employees)").fetchall()]
    if "appraisal_supervisor_username" not in emp_cols:
        db.execute("ALTER TABLE employees ADD COLUMN appraisal_supervisor_username TEXT")
        applied.append("employees.appraisal_supervisor_username")
    if "medical_claim_limit" not in emp_cols:
        db.execute("ALTER TABLE employees ADD COLUMN medical_claim_limit REAL DEFAULT 0")
        applied.append("employees.medical_claim_limit")
    if "ot_approval_required" not in emp_cols:
        db.execute("ALTER TABLE employees ADD COLUMN ot_approval_required TEXT NOT NULL DEFAULT 'N'")
        applied.append("employees.ot_approval_required")
    for col in ["emergency_contact_1_name", "emergency_contact_1_phone", "emergency_contact_1_relationship",
                "emergency_contact_2_name", "emergency_contact_2_phone", "emergency_contact_2_relationship",
                "hr_username"]:
        if col not in emp_cols:
            db.execute(f"ALTER TABLE employees ADD COLUMN {col} TEXT")
            applied.append(f"employees.{col}")

    existing_tables = {r[0] for r in db.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    if "appraisals" not in existing_tables:
        db.execute("""CREATE TABLE appraisals (
            id INTEGER PRIMARY KEY AUTOINCREMENT, emp_id TEXT NOT NULL REFERENCES employees(emp_id),
            purpose TEXT NOT NULL DEFAULT 'Assessment', appraisal_date TEXT NOT NULL,
            ratings_json TEXT NOT NULL, total_score REAL NOT NULL, max_score REAL NOT NULL,
            percentage REAL NOT NULL, overall_band TEXT NOT NULL,
            rec_advancement TEXT NOT NULL DEFAULT 'N', rec_not_yet_ready TEXT NOT NULL DEFAULT 'N',
            rec_better_suited TEXT NOT NULL DEFAULT 'N', rec_training_required TEXT NOT NULL DEFAULT 'N',
            comments TEXT, status TEXT NOT NULL DEFAULT 'Draft', supervisor_username TEXT NOT NULL,
            submitted_at TEXT, current_salary REAL, increment_amount REAL DEFAULT 0, new_salary REAL)""")
        applied.append("table: appraisals")
    if "appraisals" in existing_tables:
        appraisal_cols = [r[1] for r in db.execute("PRAGMA table_info(appraisals)").fetchall()]
        for col, decl in [("current_salary", "REAL"), ("increment_amount", "REAL DEFAULT 0"), ("new_salary", "REAL"),
                           ("hr_viewed_at", "TEXT")]:
            if col not in appraisal_cols:
                db.execute(f"ALTER TABLE appraisals ADD COLUMN {col} {decl}")
                applied.append(f"appraisals.{col}")
    if "payroll_runs" in existing_tables:
        payroll_runs_cols = [r[1] for r in db.execute("PRAGMA table_info(payroll_runs)").fetchall()]
        for col, decl in [("pcb_override", "REAL"), ("pcb_override_reason", "TEXT")]:
            if col not in payroll_runs_cols:
                db.execute(f"ALTER TABLE payroll_runs ADD COLUMN {col} {decl}")
                applied.append(f"payroll_runs.{col}")
    if "medical_claims" not in existing_tables:
        db.execute("""CREATE TABLE medical_claims (
            id INTEGER PRIMARY KEY AUTOINCREMENT, emp_id TEXT NOT NULL REFERENCES employees(emp_id),
            claim_date TEXT NOT NULL, amount REAL NOT NULL, clinic_name TEXT, description TEXT,
            supporting_doc_original TEXT, supporting_doc_stored TEXT, status TEXT NOT NULL DEFAULT 'Pending',
            submitted_at TEXT NOT NULL, reviewed_by TEXT, reviewed_at TEXT, review_notes TEXT)""")
        applied.append("table: medical_claims")
    if "profile_update_log" not in existing_tables:
        db.execute("""CREATE TABLE profile_update_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT, emp_id TEXT NOT NULL REFERENCES employees(emp_id),
            updated_at TEXT NOT NULL, hr_viewed_at TEXT)""")
        applied.append("table: profile_update_log")
    if "business_trips" in existing_tables:
        bt_cols = [r[1] for r in db.execute("PRAGMA table_info(business_trips)").fetchall()]
        if "notice_type" not in bt_cols:
            db.execute("ALTER TABLE business_trips ADD COLUMN notice_type TEXT NOT NULL DEFAULT 'Business Trip'")
            applied.append("business_trips.notice_type")
        for col in ("supporting_doc_original", "supporting_doc_stored"):
            if col not in bt_cols:
                db.execute(f"ALTER TABLE business_trips ADD COLUMN {col} TEXT")
                applied.append(f"business_trips.{col}")
    if "attendance_daily" not in existing_tables:
        db.execute("""CREATE TABLE attendance_daily (
            id INTEGER PRIMARY KEY AUTOINCREMENT, emp_id TEXT NOT NULL REFERENCES employees(emp_id),
            date TEXT NOT NULL, day_type TEXT NOT NULL DEFAULT 'WORKED',
            time_in TEXT, time_out TEXT, meal_allowance_flag TEXT NOT NULL DEFAULT 'N',
            ot_hours_1_5 REAL DEFAULT 0, ot_hours_2_0 REAL DEFAULT 0, ot_hours_3_0 REAL DEFAULT 0,
            ot_reason TEXT, UNIQUE (emp_id, date))""")
        applied.append("table: attendance_daily")
    if "ot_claims" not in existing_tables:
        db.execute("""CREATE TABLE ot_claims (
            id INTEGER PRIMARY KEY AUTOINCREMENT, emp_id TEXT NOT NULL REFERENCES employees(emp_id),
            claim_date TEXT NOT NULL, ot_hours_1_5 REAL NOT NULL DEFAULT 0,
            ot_hours_2_0 REAL NOT NULL DEFAULT 0, ot_hours_3_0 REAL NOT NULL DEFAULT 0,
            reason TEXT, status TEXT NOT NULL DEFAULT 'Pending', submitted_by TEXT NOT NULL,
            submitted_at TEXT NOT NULL, reviewed_by TEXT, reviewed_at TEXT, review_notes TEXT)""")
        applied.append("table: ot_claims")

    db.execute("UPDATE hr_users SET can_approve_leave='Y', can_approve_appraisal='Y' WHERE username='kee'")
    yang_password = request.form.get("yang_password", "")
    if not db.execute("SELECT 1 FROM hr_users WHERE username='yang'").fetchone():
        db.execute(
            "INSERT INTO hr_users (username, password_hash, full_name, created_at, role, can_approve_leave, can_approve_appraisal, can_approve_ot) VALUES (?,?,?,?,?,?,?,?)",
            ("yang", generate_password_hash(yang_password), "Yang Hui",
             datetime.datetime.now().isoformat(timespec="seconds"), "approver", "N", "Y", "Y"),
        )
        applied.append("hr_users: yang (created)")
    elif yang_password and request.form.get("reset_yang_password") == "1":
        db.execute("UPDATE hr_users SET password_hash=? WHERE username='yang'", (generate_password_hash(yang_password),))
        applied.append("hr_users: yang (password reset)")
    if db.execute("SELECT 1 FROM hr_users WHERE username='yang' AND can_approve_ot != 'Y'").fetchone():
        db.execute("UPDATE hr_users SET can_approve_ot='Y' WHERE username='yang'")
        applied.append("hr_users: yang (can_approve_ot)")
    db.commit()
    return "OK - applied: " + (", ".join(applied) if applied else "(nothing new, already up to date)"), 200


@app.route("/hr/seed-ot-claims", methods=["POST"])
def hr_seed_ot_claims():
    """One-time helper: flags an employee ot_approval_required='Y' and
    inserts one or more Pending OT claims on their behalf, so OT that
    predates the OT Claim approval flow can be re-recorded properly
    through it (for review/approval by the employee's OT approver)
    without a live HR session. Does not touch attendance/payroll itself -
    that only happens once a claim is actually approved, same as any
    other OT claim. Same RESTORE_TOKEN gate as the other one-time routes;
    safe to re-run (each call just inserts more Pending claims).

    Expected JSON body: {"emp_id": "I001", "claims": [
        {"claim_date": "2026-08-17", "ot_hours_1_5": 1.5, "ot_hours_2_0": 0,
         "ot_hours_3_0": 0, "reason": "..."}, ...]}
    """
    token = os.environ.get("RESTORE_TOKEN")
    if not token or request.form.get("token") != token:
        abort(404)
    payload = json.loads(request.files["data"].read())
    db = get_db()
    emp_id = payload["emp_id"]
    if db.execute("SELECT 1 FROM employees WHERE emp_id=?", (emp_id,)).fetchone() is None:
        return f"Refused: unknown emp_id {emp_id}", 400
    db.execute("UPDATE employees SET ot_approval_required='Y' WHERE emp_id=?", (emp_id,))
    now = datetime.datetime.now().isoformat(timespec="seconds")
    written = 0
    for claim in payload.get("claims", []):
        db.execute(
            """INSERT INTO ot_claims (emp_id, claim_date, ot_hours_1_5, ot_hours_2_0, ot_hours_3_0,
                   reason, status, submitted_by, submitted_at)
               VALUES (?,?,?,?,?,?,'Pending','HR',?)""",
            (emp_id, claim["claim_date"], claim.get("ot_hours_1_5", 0) or 0,
             claim.get("ot_hours_2_0", 0) or 0, claim.get("ot_hours_3_0", 0) or 0,
             claim.get("reason"), now),
        )
        written += 1
    db.commit()
    return f"OK - flagged {emp_id} ot_approval_required=Y, wrote {written} pending OT claim(s)", 200


@app.route("/hr/ot-claims-cleanup", methods=["POST"])
def hr_ot_claims_cleanup():
    """One-time helper to reject specific OT claim IDs directly (e.g. to
    remove accidental duplicates from a scripted seed) without a live HR
    session. Only touches rows still Pending, so it can't undo a decision
    someone already made through the normal approval UI. Same
    RESTORE_TOKEN gate as the other one-time routes."""
    token = os.environ.get("RESTORE_TOKEN")
    if not token or request.form.get("token") != token:
        abort(404)
    ids = [int(x) for x in request.form.get("ids", "").split(",") if x.strip()]
    notes = request.form.get("review_notes") or None
    db = get_db()
    now = datetime.datetime.now().isoformat(timespec="seconds")
    updated = 0
    for claim_id in ids:
        cur = db.execute(
            """UPDATE ot_claims SET status='Rejected', reviewed_by='HR', reviewed_at=?, review_notes=?
               WHERE id=? AND status='Pending'""",
            (now, notes, claim_id),
        )
        updated += cur.rowcount
    db.commit()
    return f"OK - rejected {updated} claim(s)", 200


@app.route("/hr/import-historical-payroll", methods=["POST"])
def hr_import_historical_payroll():
    """One-time replacement of Jan-Jul 2026 payroll_runs figures with the
    real numbers from the user's own source Excel files, for months that
    were finalized before the PCB year-to-date history fix existed and so
    had systematically wrong PCB (and, for some employees, EPF/SKBBK) -
    see the "just follow my excel" decision. Refuses anything August 2026
    or later so the accurate live-engine months can never be touched.
    Same RESTORE_TOKEN gate as the other one-time migration routes above;
    safe to re-run (upserts on the emp_id/year/month key)."""
    token = os.environ.get("RESTORE_TOKEN")
    if not token or request.form.get("token") != token:
        abort(404)
    file = request.files.get("file")
    if file is None or file.filename == "":
        return "No file uploaded", 400
    records = json.load(file)
    db = get_db()
    known_emp_ids = {r["emp_id"] for r in db.execute("SELECT emp_id FROM employees").fetchall()}
    for r in records:
        if (r["year"], r["month"]) >= (2026, 8):
            return f"Refused: record for {r['emp_id']} {r['year']}-{r['month']:02d} is Aug 2026 or later - this import is Jan-Jul 2026 only", 400
        if r["emp_id"] not in known_emp_ids:
            return f"Refused: unknown emp_id {r['emp_id']}", 400
    now = datetime.datetime.now().isoformat(timespec="seconds")
    written = []
    for r in records:
        db.execute(
            """INSERT INTO payroll_runs (
                emp_id, year, month, basic_salary, fixed_allowance, variable_allowance,
                working_days_in_month, days_worked, paid_leave_days, unpaid_days, unpaid_deduction,
                ot_hours_1_5, ot_hours_2_0, ot_hours_3_0, ot_pay_1_5, ot_pay_2_0, ot_pay_3_0,
                ot_hourly_rate, ot_pay, days_employed, prorate_factor, transport_allowance,
                meal_allowance, cewi_allowance, gross_pay, epf_employee, additional_epf_employee,
                epf_employer, socso_employee, socso_employer, eis_employee, eis_employer, pcb,
                skbbk_employee, hrd_levy_employer, other_deduction, other_deduction_desc,
                total_deductions, net_pay, finalized_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(emp_id, year, month) DO UPDATE SET
                basic_salary=excluded.basic_salary, fixed_allowance=excluded.fixed_allowance,
                variable_allowance=excluded.variable_allowance,
                working_days_in_month=excluded.working_days_in_month,
                days_worked=excluded.days_worked, paid_leave_days=excluded.paid_leave_days,
                unpaid_days=excluded.unpaid_days, unpaid_deduction=excluded.unpaid_deduction,
                ot_hours_1_5=excluded.ot_hours_1_5, ot_hours_2_0=excluded.ot_hours_2_0,
                ot_hours_3_0=excluded.ot_hours_3_0, ot_pay_1_5=excluded.ot_pay_1_5,
                ot_pay_2_0=excluded.ot_pay_2_0, ot_pay_3_0=excluded.ot_pay_3_0,
                ot_hourly_rate=excluded.ot_hourly_rate, ot_pay=excluded.ot_pay,
                days_employed=excluded.days_employed, prorate_factor=excluded.prorate_factor,
                transport_allowance=excluded.transport_allowance,
                meal_allowance=excluded.meal_allowance, cewi_allowance=excluded.cewi_allowance,
                gross_pay=excluded.gross_pay, epf_employee=excluded.epf_employee,
                additional_epf_employee=excluded.additional_epf_employee,
                epf_employer=excluded.epf_employer, socso_employee=excluded.socso_employee,
                socso_employer=excluded.socso_employer, eis_employee=excluded.eis_employee,
                eis_employer=excluded.eis_employer, pcb=excluded.pcb,
                skbbk_employee=excluded.skbbk_employee, hrd_levy_employer=excluded.hrd_levy_employer,
                other_deduction=excluded.other_deduction, other_deduction_desc=excluded.other_deduction_desc,
                total_deductions=excluded.total_deductions, net_pay=excluded.net_pay,
                finalized_at=excluded.finalized_at""",
            (
                r["emp_id"], r["year"], r["month"], r["basic_salary"], r["fixed_allowance"],
                r["variable_allowance"], r["working_days_in_month"], r["days_worked"],
                r["paid_leave_days"], r["unpaid_days"], r["unpaid_deduction"],
                r["ot_hours_1_5"], r["ot_hours_2_0"], r["ot_hours_3_0"],
                r["ot_pay_1_5"], r["ot_pay_2_0"], r["ot_pay_3_0"],
                r["ot_hourly_rate"], r["ot_pay"], r["days_employed"], r["prorate_factor"],
                r["transport_allowance"], r["meal_allowance"], r["cewi_allowance"],
                r["gross_pay"], r["epf_employee"], r["additional_epf_employee"],
                r["epf_employer"], r["socso_employee"], r["socso_employer"],
                r["eis_employee"], r["eis_employer"], r["pcb"],
                r["skbbk_employee"], r["hrd_levy_employer"], r["other_deduction"], r["other_deduction_desc"],
                r["total_deductions"], r["net_pay"], now,
            ),
        )
        db.execute(
            """INSERT INTO pcb_monthly_record (emp_id, year, month, gross_remun, epf_employee, pcb_deducted)
               VALUES (?,?,?,?,?,?)
               ON CONFLICT(emp_id, year, month) DO UPDATE SET
                 gross_remun=excluded.gross_remun, epf_employee=excluded.epf_employee,
                 pcb_deducted=excluded.pcb_deducted""",
            (r["emp_id"], r["year"], r["month"], r["gross_pay"], r["epf_employee"], r["pcb"]),
        )
        written.append(f"{r['emp_id']} {r['year']}-{r['month']:02d}")
    db.commit()
    return f"OK - imported {len(written)} records: " + ", ".join(written), 200


@app.route("/hr/backup-database")
def hr_backup_database():
    """Downloads the live payroll.db as-is - a manual safety copy, since
    the current Railway plan doesn't include automatic backups. Same
    RESTORE_TOKEN gate as the other one-time routes; read-only, makes no
    changes."""
    token = os.environ.get("RESTORE_TOKEN")
    if not token or request.args.get("token") != token:
        abort(404)
    return send_from_directory(
        DATA_DIR, "payroll.db", as_attachment=True,
        download_name=f"payroll_backup_{datetime.date.today().isoformat()}.db",
    )


@app.route("/hr/backup-uploads")
def hr_backup_uploads():
    """Downloads the uploads/ folder (photos, IC copies, receipts, other
    employee documents) as a single zip. Same RESTORE_TOKEN gate; read-only."""
    token = os.environ.get("RESTORE_TOKEN")
    if not token or request.args.get("token") != token:
        abort(404)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, _dirs, files in os.walk(UPLOAD_DIR):
            for fname in files:
                full_path = os.path.join(root, fname)
                zf.write(full_path, os.path.relpath(full_path, UPLOAD_DIR))
    buf.seek(0)
    filename = f"uploads_backup_{datetime.date.today().isoformat()}.zip"
    return Response(
        buf.getvalue(),
        mimetype="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


ADJUSTABLE_ATTENDANCE_FIELDS = {
    "days_worked", "al_days", "mc_days", "hl_days", "ul_days", "other_paid_leave",
    "ph_days", "off_days", "rest_days", "absent_days", "working_days_in_month",
}


@app.route("/hr/adjust-attendance", methods=["POST"])
def hr_adjust_attendance():
    """One-off correction to a single attendance_monthly field for one
    employee/month (e.g. adding MC days from a paper leave slip), without
    disturbing the rest of that month's figures the way a full
    daily-attendance resync would if only some days were known. Adds
    `delta` to whatever the field currently holds (0 if no row exists
    yet). Same RESTORE_TOKEN gate as the other one-time routes."""
    token = os.environ.get("RESTORE_TOKEN")
    if not token or request.form.get("token") != token:
        abort(404)
    emp_id = request.form.get("emp_id")
    year = request.form.get("year", type=int)
    month = request.form.get("month", type=int)
    field = request.form.get("field")
    delta = request.form.get("delta", type=float)
    if field not in ADJUSTABLE_ATTENDANCE_FIELDS:
        return f"Refused: field must be one of {sorted(ADJUSTABLE_ATTENDANCE_FIELDS)}", 400
    db = get_db()
    if db.execute("SELECT 1 FROM employees WHERE emp_id=?", (emp_id,)).fetchone() is None:
        return f"Refused: unknown emp_id {emp_id}", 400
    db.execute(
        f"""INSERT INTO attendance_monthly (emp_id, year, month, {field}) VALUES (?,?,?,?)
           ON CONFLICT(emp_id, year, month) DO UPDATE SET {field} = COALESCE({field}, 0) + ?""",
        (emp_id, year, month, delta, delta),
    )
    db.commit()
    new_val = db.execute(
        f"SELECT {field} FROM attendance_monthly WHERE emp_id=? AND year=? AND month=?",
        (emp_id, year, month),
    ).fetchone()[0]
    return f"OK - {emp_id} {year}-{month:02d} {field} adjusted by {delta:+g}, now {new_val}", 200


@app.route("/hr/seed-attendance-daily", methods=["POST"])
def hr_seed_attendance_daily():
    """One-time helper: accepts a JSON payload of daily attendance rows for
    one or more employees and writes them via the same path the Daily
    Attendance page itself uses (upsert into attendance_daily, then
    recompute attendance_monthly), so a real attendance sheet can be keyed
    in directly without an HR session. Same RESTORE_TOKEN gate as the
    other one-time routes; safe to re-run (upserts on emp_id/date).

    Expected JSON body: {"EMP_ID": {"YYYY-MM-DD": {"day_type": "WORKED",
    "time_in": "08:30", "time_out": "17:30", "meal": "Y",
    "ot_hours_1_5": 0, "ot_hours_2_0": 0, "ot_hours_3_0": 0,
    "ot_reason": null}, ...}, ...}
    """
    token = os.environ.get("RESTORE_TOKEN")
    if not token or request.form.get("token") != token:
        abort(404)
    payload = json.loads(request.files["data"].read())
    db = get_db()
    known_emp_ids = {r["emp_id"] for r in db.execute("SELECT emp_id FROM employees").fetchall()}
    months_touched = set()
    written = 0
    for emp_id, day_rows in payload.items():
        if emp_id not in known_emp_ids:
            return f"Refused: unknown emp_id {emp_id}", 400
        for date_str, d in day_rows.items():
            year, month, _day = (int(p) for p in date_str.split("-"))
            db.execute(
                """INSERT INTO attendance_daily (
                       emp_id, date, day_type, time_in, time_out, meal_allowance_flag,
                       ot_hours_1_5, ot_hours_2_0, ot_hours_3_0, ot_reason
                   ) VALUES (?,?,?,?,?,?,?,?,?,?)
                   ON CONFLICT(emp_id, date) DO UPDATE SET
                     day_type=excluded.day_type, time_in=excluded.time_in, time_out=excluded.time_out,
                     meal_allowance_flag=excluded.meal_allowance_flag,
                     ot_hours_1_5=excluded.ot_hours_1_5, ot_hours_2_0=excluded.ot_hours_2_0,
                     ot_hours_3_0=excluded.ot_hours_3_0, ot_reason=excluded.ot_reason""",
                (emp_id, date_str, d.get("day_type", "WORKED"), d.get("time_in") or None,
                 d.get("time_out") or None, d.get("meal", "N"),
                 d.get("ot_hours_1_5", 0) or 0, d.get("ot_hours_2_0", 0) or 0,
                 d.get("ot_hours_3_0", 0) or 0, d.get("ot_reason") or None),
            )
            written += 1
            months_touched.add((emp_id, year, month))
    for emp_id, year, month in months_touched:
        _sync_daily_to_monthly(db, emp_id, year, month)
    db.commit()
    return f"OK - wrote {written} daily rows across {len(months_touched)} employee-month(s)", 200


@app.route("/hr/bulk-set-medical-claim-limit", methods=["POST"])
def hr_bulk_set_medical_claim_limit():
    """One-time fix: sets the company-standard RM500/year medical claim
    limit for every employee who doesn't already have a custom (nonzero)
    limit set. Same RESTORE_TOKEN gate as the other one-time routes above;
    safe to re-run since it only ever touches rows still at 0/NULL."""
    token = os.environ.get("RESTORE_TOKEN")
    if not token or request.form.get("token") != token:
        abort(404)
    db = get_db()
    cur = db.execute(
        "UPDATE employees SET medical_claim_limit=500 WHERE medical_claim_limit IS NULL OR medical_claim_limit=0"
    )
    db.commit()
    return f"OK - updated {cur.rowcount} employee(s) to RM500/year", 200


if __name__ == "__main__":
    app.run(debug=True, port=5000)
