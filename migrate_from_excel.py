# -*- coding: utf-8 -*-
"""
One-time import from Payroll_Master_2026_9.xlsm into payroll.db.

Column letters below are taken directly from the formulas in the 'Payroll'
and 'PCB Inputs & YTD' sheets (e.g. P4 = IF(BA4="Y",AU4,0)+... tells us BA is
the Position Allowance flag and AU is its amount) rather than from the row-3
header labels, since several flag columns are unlabeled in the sheet itself.
"""
import datetime
import os
import sqlite3
import sys

import openpyxl

SRC = r"E:\HRMS TEST\Payroll_Master_2026_9.xlsm"
DB_PATH = os.path.join(os.path.dirname(__file__), "payroll.db")
SCHEMA_PATH = os.path.join(os.path.dirname(__file__), "schema.sql")

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul",
          "Aug", "Sep", "Oct", "Nov", "Dec"]


def to_iso_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def to_num(v, default=0):
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def to_str(v):
    if v is None:
        return None
    return str(v).strip()


def main():
    if os.path.exists(DB_PATH):
        print(f"Refusing to overwrite existing {DB_PATH}. Delete it first if you want to re-import.")
        sys.exit(1)

    wb = openpyxl.load_workbook(SRC, data_only=True)

    conn = sqlite3.connect(DB_PATH)
    with open(SCHEMA_PATH, "r", encoding="utf-8") as f:
        conn.executescript(f.read())

    # ---------- Employee Master ----------
    ws = wb["Employee Master"]
    emp_ids = []
    row = 4
    empty_streak = 0
    while empty_streak < 3:
        emp_id = to_str(ws[f"A{row}"].value)
        if not emp_id:
            empty_streak += 1
            row += 1
            continue
        empty_streak = 0
        emp_ids.append(emp_id)

        conn.execute(
            """INSERT INTO employees (
                emp_id, full_name, ic_passport_no, date_of_birth, race, religion,
                marital_status, department, position, date_joined, status,
                holiday_state, basic_salary, working_days_week, working_hours_day,
                standard_start, standard_end, epf_no, socso_no, bank_name,
                bank_account_no, annual_leave_entitlement, mc_entitlement,
                hospitalisation_leave_entitlement, lunch_start, lunch_end,
                passport_expiry, work_permit_expiry, probation_end_date,
                retirement_date, last_working_day, skbbk_flag,
                transport_allowance, transport_allowance_flag,
                meal_allowance_rate, meal_allowance_flag,
                position_allowance, position_allowance_flag,
                cewi_rate, cewi_flag,
                training_incentive, training_incentive_flag,
                oversea_incentive, oversea_incentive_flag
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                emp_id,
                to_str(ws[f"B{row}"].value),
                to_str(ws[f"C{row}"].value),
                to_iso_date(ws[f"D{row}"].value),
                to_str(ws[f"F{row}"].value),
                to_str(ws[f"G{row}"].value),
                to_str(ws[f"H{row}"].value),
                to_str(ws[f"I{row}"].value),
                to_str(ws[f"J{row}"].value),
                to_iso_date(ws[f"K{row}"].value),
                to_str(ws[f"M{row}"].value),
                to_str(ws[f"N{row}"].value),
                to_num(ws[f"O{row}"].value),
                to_num(ws[f"Q{row}"].value, None),
                to_num(ws[f"R{row}"].value, None),
                to_str(ws[f"S{row}"].value),
                to_str(ws[f"T{row}"].value),
                to_str(ws[f"U{row}"].value),
                to_str(ws[f"V{row}"].value),
                to_str(ws[f"W{row}"].value),
                to_str(ws[f"X{row}"].value),
                to_num(ws[f"Y{row}"].value, None),
                to_num(ws[f"Z{row}"].value, None),
                to_num(ws[f"AA{row}"].value, None),
                to_str(ws[f"AJ{row}"].value),
                to_str(ws[f"AK{row}"].value),
                to_iso_date(ws[f"AL{row}"].value),
                to_iso_date(ws[f"AM{row}"].value),
                to_iso_date(ws[f"AN{row}"].value),
                to_iso_date(ws[f"AO{row}"].value),
                to_iso_date(ws[f"AP{row}"].value),
                to_str(ws[f"AR{row}"].value),
                to_num(ws[f"AS{row}"].value),
                to_str(ws[f"AY{row}"].value) or "N",
                to_num(ws[f"AT{row}"].value),
                to_str(ws[f"AZ{row}"].value) or "N",
                to_num(ws[f"AU{row}"].value),
                to_str(ws[f"BA{row}"].value) or "N",
                to_num(ws[f"AV{row}"].value),
                to_str(ws[f"BB{row}"].value) or "N",
                to_num(ws[f"AW{row}"].value),
                to_str(ws[f"BC{row}"].value) or "N",
                to_num(ws[f"AX{row}"].value),
                to_str(ws[f"BD{row}"].value) or "N",
            ),
        )
        row += 1
    print(f"Employees imported: {len(emp_ids)}")

    # ---------- Public Holidays ----------
    ws = wb["Public Holidays"]
    row = 3
    ph_count = 0
    while True:
        date_val = ws[f"A{row}"].value
        if date_val is None:
            break
        conn.execute(
            "INSERT INTO public_holidays (date, day, name, remarks) VALUES (?,?,?,?)",
            (to_iso_date(date_val), to_str(ws[f"B{row}"].value),
             to_str(ws[f"C{row}"].value), to_str(ws[f"D{row}"].value)),
        )
        ph_count += 1
        row += 1
    print(f"Public holidays imported: {ph_count}")

    # ---------- Monthly attendance ----------
    att_count = 0
    for month_idx, month_name in enumerate(MONTHS, start=1):
        if month_name not in wb.sheetnames:
            continue
        ws = wb[month_name]
        row = 4
        empty_streak = 0
        while empty_streak < 3:
            emp_id = to_str(ws[f"A{row}"].value)
            if not emp_id:
                empty_streak += 1
                row += 1
                continue
            empty_streak = 0
            if emp_id not in emp_ids:
                row += 1
                continue
            # skip rows with no attendance data at all (never filled in)
            if all(ws[f"{c}{row}"].value in (None, "") for c in
                   "CDEFGHIJKLMNOPQR"):
                row += 1
                continue
            conn.execute(
                """INSERT INTO attendance_monthly (
                    emp_id, year, month, days_worked, al_days, mc_days, hl_days,
                    ul_days, other_paid_leave, ph_days, off_days, rest_days,
                    absent_days, working_days_in_month, late_in_count,
                    early_out_count, lunch_late_count, ot_hours_approved,
                    meal_eligible_days
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    emp_id, 2026, month_idx,
                    to_num(ws[f"C{row}"].value), to_num(ws[f"D{row}"].value),
                    to_num(ws[f"E{row}"].value), to_num(ws[f"F{row}"].value),
                    to_num(ws[f"G{row}"].value), to_num(ws[f"H{row}"].value),
                    to_num(ws[f"I{row}"].value), to_num(ws[f"J{row}"].value),
                    to_num(ws[f"K{row}"].value), to_num(ws[f"L{row}"].value),
                    to_num(ws[f"M{row}"].value), to_num(ws[f"N{row}"].value),
                    to_num(ws[f"O{row}"].value), to_num(ws[f"P{row}"].value),
                    to_num(ws[f"Q{row}"].value), to_num(ws[f"R{row}"].value),
                ),
            )
            att_count += 1
            row += 1
    print(f"Attendance-month rows imported: {att_count}")

    # ---------- Leave Types ----------
    ws = wb["Leave Types"]
    row = 3
    while True:
        code = to_str(ws[f"A{row}"].value)
        if not code:
            break
        conn.execute(
            "INSERT INTO leave_types (code, description, paid, notes) VALUES (?,?,?,?)",
            (code, to_str(ws[f"B{row}"].value), to_str(ws[f"C{row}"].value),
             to_str(ws[f"D{row}"].value)),
        )
        row += 1

    row = 19
    while True:
        v = ws[f"A{row}"].value
        if v is None or v == "":
            break
        conn.execute(
            "INSERT INTO leave_entitlement_by_yos (min_years_service, annual_leave_days, mc_days, notes) VALUES (?,?,?,?)",
            (to_num(v), int(to_num(ws[f"B{row}"].value)), int(to_num(ws[f"C{row}"].value)),
             to_str(ws[f"D{row}"].value)),
        )
        row += 1
    print("Leave types and entitlement table imported")

    # ---------- EPF Table ----------
    ws = wb["EPF Table"]
    row = 4
    epf_count = 0
    while True:
        v = ws[f"A{row}"].value
        if v is None or v == "":
            break
        conn.execute(
            """INSERT INTO epf_table (
                wage_lower_bound, bracket_label, part_a_under60_employer,
                part_a_under60_employee, part_e_60plus_employer,
                part_e_60plus_employee
            ) VALUES (?,?,?,?,?,?)""",
            (to_num(v), to_str(ws[f"B{row}"].value), to_num(ws[f"C{row}"].value),
             to_num(ws[f"D{row}"].value), to_num(ws[f"F{row}"].value),
             to_num(ws[f"G{row}"].value)),
        )
        epf_count += 1
        row += 1
    print(f"EPF table brackets imported: {epf_count}")

    # ---------- SOCSO-SKBBK Table ----------
    ws = wb["SOCSO-SKBBK Table"]
    row = 4
    socso_count = 0
    while True:
        v = ws[f"A{row}"].value
        if v is None or v == "":
            break
        conn.execute(
            """INSERT INTO socso_table (
                wage_lower_bound, bracket_label, cat1_employer,
                cat1_employee_invalidity, cat1_employee_skbbk,
                cat2_employer, cat2_employee_skbbk
            ) VALUES (?,?,?,?,?,?,?)""",
            (to_num(v), to_str(ws[f"B{row}"].value), to_num(ws[f"C{row}"].value),
             to_num(ws[f"D{row}"].value), to_num(ws[f"E{row}"].value),
             to_num(ws[f"H{row}"].value), to_num(ws[f"I{row}"].value)),
        )
        socso_count += 1
        row += 1
    print(f"SOCSO table brackets imported: {socso_count}")

    # ---------- PCB Tax Rate Table ----------
    ws = wb["PCB Tax Rate Table"]
    const_map = {
        "individual_relief": "B5",
        "spouse_relief": "B6",
        "child_relief_full": "B7",
        "child_relief_half": "B8",
        "epf_relief_cap": "B9",
        "socso_eis_relief_cap": "B10",
        "rebate_threshold": "B11",
        "rebate_amount": "B12",
    }
    for key, cell in const_map.items():
        conn.execute("INSERT INTO pcb_constants (key, value) VALUES (?,?)",
                     (key, to_num(ws[cell].value)))

    row = 16
    bracket_count = 0
    while True:
        v = ws[f"A{row}"].value
        if v is None or v == "":
            break
        conn.execute(
            "INSERT INTO pcb_tax_brackets (income_from, rate, base_tax) VALUES (?,?,?)",
            (to_num(v), to_num(ws[f"B{row}"].value), to_num(ws[f"C{row}"].value)),
        )
        bracket_count += 1
        row += 1
    print(f"PCB tax brackets imported: {bracket_count}, constants: {len(const_map)}")

    # ---------- Tax Profile (PCB Inputs & YTD, rows 6..) ----------
    ws = wb["PCB Inputs & YTD"]
    row = 6
    tp_count = 0
    for emp_id in emp_ids:
        row_emp_id = to_str(ws[f"A{row}"].value)
        if row_emp_id != emp_id:
            # Fall back: scan for it (defensive, shouldn't normally trigger
            # since Tax Profile rows mirror Employee Master rows 1:1)
            found = None
            for r in range(6, 6 + len(emp_ids) + 5):
                if to_str(ws[f"A{r}"].value) == emp_id:
                    found = r
                    break
            if found is None:
                row += 1
                continue
            row_use = found
        else:
            row_use = row
        conn.execute(
            """INSERT INTO tax_profile (
                emp_id, tax_category, children_full_relief, children_half_relief,
                tp1_submitted, tp1_date, zakat_paid_ytd
            ) VALUES (?,?,?,?,?,?,?)""",
            (
                emp_id,
                to_str(ws[f"C{row_use}"].value) or "Single",
                int(to_num(ws[f"D{row_use}"].value)),
                int(to_num(ws[f"E{row_use}"].value)),
                to_str(ws[f"F{row_use}"].value) or "",
                to_iso_date(ws[f"G{row_use}"].value),
                to_num(ws[f"H{row_use}"].value),
            ),
        )
        tp_count += 1
        row += 1
    print(f"Tax profiles imported: {tp_count}")

    # ---------- Payroll settings ----------
    ws = wb["Payroll"]
    settings = {
        "eis_wage_ceiling": ws["AA2"].value,
        "eis_employee_pct": ws["AA4"].value,
        "eis_employer_pct": ws["AA5"].value,
        "ot_rate_multiplier": ws["AA6"].value,
        "hrd_levy_registered": ws["AA10"].value,
        "hrd_levy_rate": ws["AA11"].value,
    }
    for key, value in settings.items():
        conn.execute("INSERT INTO payroll_settings (key, value) VALUES (?,?)",
                     (key, str(value)))
    print(f"Payroll settings imported: {settings}")

    conn.commit()
    conn.close()
    print(f"\nDone. Database written to {DB_PATH}")


if __name__ == "__main__":
    main()
